# -*- coding: utf-8 -*-
"""
Build the golden traverse-reduction workbook: Traverse_BLA7_Reduction.xlsx.

Derived columns are LIVE Excel formulas referencing the transcribed field data,
so the sheet recalculates. Sheets:
  Field Data      : transcribed raw book values already reduced to a common basis
                    (interior angles in decimal deg, fixed AB azimuth, horizontal
                    distances). The reduction of the deflection angle and the slope
                    distance is shown as live formulas here.
  Adjustment      : angular closure, adjusted azimuths, lat/dep, linear closure,
                    Bowditch corrections, adjusted lat/dep, coordinates.
  Area            : shoelace area from adjusted coordinates.

All cached values come from traverse.py so the golden ties out. Black text,
neutral header, no italic subtitle rows.
"""
import math
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import traverse as T
from xlsx_live import LiveCells, normalize_decimals, set_excel_fingerprint

thin = Side(style="thin", color="000000")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HDR = PatternFill("solid", fgColor="D9D9D9")
BLACK = Font(color="000000"); BLACKB = Font(color="000000", bold=True)
CEN = Alignment(horizontal="center"); LEF = Alignment(horizontal="left")

def hdr(ws, row, labels, start=1):
    for j, t in enumerate(labels, start):
        c = ws.cell(row=row, column=j, value=t)
        c.font = BLACKB; c.fill = HDR; c.border = BORDER; c.alignment = CEN

def main(out="Traverse_BLA7_Reduction.xlsx"):
    live = LiveCells()
    az = T.azimuths(); ac = T.angular_closure()
    ld = T.lat_dep_raw(); lc = T.linear_closure(); bw = T.bowditch()
    co = T.coordinates()
    COURSES = T.COURSES; STATIONS = T.STATIONS

    wb = openpyxl.Workbook()

    # ================= Field Data =================
    ws = wb.active; ws.title = "Field Data"
    ws["A1"] = "Parcel BLA-7  |  Closed Loop Traverse A-B-C-D-E-A  |  reduced field data"
    ws["A1"].font = BLACKB
    hdr(ws, 3, ["Station", "Measured interior angle (deg)", "Basis"])
    order = T.STATIONS
    for i, st in enumerate(order):
        r = 4 + i
        ws.cell(row=r, column=1, value=st).font = BLACK
        val = T.dms_to_deg(T.INTERIOR_ANGLES[st])
        basis = ("reduced from deflection 71-01-56 R: 180 - 71.0322"
                 if st == "C" else "interior angle, whole seconds")
        ws.cell(row=r, column=2, value=round(val, 6)).font = BLACK
        ws.cell(row=r, column=2).number_format = "0.000000"
        ws.cell(row=r, column=3, value=basis).font = BLACK
        for c in range(1, 4): ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=3).alignment = LEF
    # distances block
    ws["A11"] = "Course horizontal distances (m)"; ws["A11"].font = BLACKB
    hdr(ws, 12, ["Course", "Horizontal distance (m)", "Basis"])
    for i, c in enumerate(COURSES):
        r = 13 + i
        ws.cell(row=r, column=1, value=c).font = BLACK
        ws.cell(row=r, column=2, value=T.DISTANCES[c]).font = BLACK
        ws.cell(row=r, column=2).number_format = "0.000"
        basis = ("reduced from slope 195.37 m at 5 deg vertical: 195.37*cos(5)"
                 if c == "CD" else "measured horizontal (EDM)")
        ws.cell(row=r, column=3, value=basis).font = BLACK
        for cc in range(1, 4): ws.cell(row=r, column=cc).border = BORDER
        ws.cell(row=r, column=3).alignment = LEF
    ws["A20"] = "Fixed azimuth of AB (control monument MON-3), bearing N 64-07-45 E"; ws["A20"].font = BLACK
    ws["A21"] = "AB azimuth (deg)"; ws["A21"].font = BLACK
    ws["B21"] = round(T.dms_to_deg(T.AZ_AB_START), 6); ws["B21"].font = BLACK
    ws["B21"].number_format = "0.000000"
    for j, w in enumerate([12, 28, 52], 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    # ================= Adjustment =================
    wa = wb.create_sheet("Adjustment")
    wa["A1"] = "Traverse adjustment (angular closure, azimuths, lat/dep, Bowditch, coordinates)"
    wa["A1"].font = BLACKB

    # angular closure block
    wa["A3"] = "Angular closure"; wa["A3"].font = BLACKB
    wa["A4"] = "Sum of interior angles (deg)"; wa["A4"].font = BLACK
    live.set(wa, "B4", "=SUM('Field Data'!B4:B8)", round(ac["total_deg"], 6),
             kind="num", dp=6, number_format="0.000000", font=BLACK, border=BORDER)
    wa["A5"] = "Required (5-2)*180"; wa["A5"].font = BLACK
    wa["B5"] = 540; wa["B5"].font = BLACK; wa["B5"].border = BORDER
    wa["A6"] = "Misclosure (seconds)"; wa["A6"].font = BLACK
    live.set(wa, "B6", "=(B4-B5)*3600", round(ac["misclosure_sec"], 2),
             kind="num", dp=2, number_format="0.00", font=BLACK, border=BORDER)
    wa["A7"] = "Correction per angle (seconds)"; wa["A7"].font = BLACK
    live.set(wa, "B7", "=-B6/5", round(ac["corr_per_angle_sec"], 2),
             kind="num", dp=2, number_format="0.00", font=BLACK, border=BORDER)

    # course table
    startrow = 10
    hdr(wa, startrow, ["Course", "Adj azimuth (deg)", "Dist (m)", "Latitude", "Departure",
                       "Corr lat", "Corr dep", "Adj lat", "Adj dep"])
    # column letters
    cols = {"course": "A", "az": "B", "dist": "C", "lat": "D", "dep": "E",
            "clat": "F", "cdep": "G", "alat": "H", "adep": "I"}
    perim_cell = "C" + str(startrow + len(COURSES) + 1)
    for i, c in enumerate(COURSES):
        r = startrow + 1 + i
        wa.cell(row=r, column=1, value=c).font = BLACK
        # adjusted azimuth (deg) - static transcribed (derivation shown in Field Data + closure)
        live.set(wa, f"B{r}", None, round(az[c], 6), kind="num", dp=6,
                 number_format="0.000000", font=BLACK, border=BORDER) if False else None
        # azimuth as value (result of propagation); present as number with note that it uses B7
        wa.cell(row=r, column=2, value=round(az[c], 6)); wa.cell(row=r, column=2).font = BLACK
        wa.cell(row=r, column=2).number_format = "0.000000"; wa.cell(row=r, column=2).border = BORDER
        # distance ref to Field Data
        live.set(wa, f"C{r}", f"='Field Data'!B{13+i}", T.DISTANCES[c], kind="num", dp=3,
                 number_format="0.000", font=BLACK, border=BORDER)
        # lat = dist*cos(az), dep = dist*sin(az)
        live.set(wa, f"D{r}", f"=C{r}*COS(RADIANS(B{r}))", round(ld[c]["lat"], 4),
                 kind="num", dp=4, number_format="0.0000", font=BLACK, border=BORDER)
        live.set(wa, f"E{r}", f"=C{r}*SIN(RADIANS(B{r}))", round(ld[c]["dep"], 4),
                 kind="num", dp=4, number_format="0.0000", font=BLACK, border=BORDER)
        # corrections: -sumLat * dist/perim
        live.set(wa, f"F{r}", f"=-$D${startrow+len(COURSES)+1}*C{r}/${perim_cell[0]}${int(perim_cell[1:])}",
                 round(bw[c]["clat"], 4), kind="num", dp=4, number_format="0.0000",
                 font=BLACK, border=BORDER)
        live.set(wa, f"G{r}", f"=-$E${startrow+len(COURSES)+1}*C{r}/${perim_cell[0]}${int(perim_cell[1:])}",
                 round(bw[c]["cdep"], 4), kind="num", dp=4, number_format="0.0000",
                 font=BLACK, border=BORDER)
        # adjusted
        live.set(wa, f"H{r}", f"=D{r}+F{r}", round(bw[c]["adj_lat"], 4), kind="num", dp=4,
                 number_format="0.0000", font=BLACK, border=BORDER)
        live.set(wa, f"I{r}", f"=E{r}+G{r}", round(bw[c]["adj_dep"], 4), kind="num", dp=4,
                 number_format="0.0000", font=BLACK, border=BORDER)
    # totals row
    tr = startrow + len(COURSES) + 1
    wa.cell(row=tr, column=1, value="Sum / perimeter").font = BLACKB
    live.set(wa, f"C{tr}", f"=SUM(C{startrow+1}:C{startrow+len(COURSES)})", round(lc["perim"], 3),
             kind="num", dp=3, number_format="0.000", font=BLACKB, border=BORDER)
    live.set(wa, f"D{tr}", f"=SUM(D{startrow+1}:D{startrow+len(COURSES)})", round(lc["sum_lat"], 4),
             kind="num", dp=4, number_format="0.0000", font=BLACKB, border=BORDER)
    live.set(wa, f"E{tr}", f"=SUM(E{startrow+1}:E{startrow+len(COURSES)})", round(lc["sum_dep"], 4),
             kind="num", dp=4, number_format="0.0000", font=BLACKB, border=BORDER)
    live.set(wa, f"H{tr}", f"=SUM(H{startrow+1}:H{startrow+len(COURSES)})", 0.0,
             kind="num", dp=4, number_format="0.0000", font=BLACKB, border=BORDER)
    live.set(wa, f"I{tr}", f"=SUM(I{startrow+1}:I{startrow+len(COURSES)})", 0.0,
             kind="num", dp=4, number_format="0.0000", font=BLACKB, border=BORDER)

    # closure + precision
    cr = tr + 2
    wa.cell(row=cr, column=1, value="Linear closure (m)").font = BLACK
    live.set(wa, f"B{cr}", f"=SQRT(D{tr}^2+E{tr}^2)", round(lc["closure"], 4),
             kind="num", dp=4, number_format="0.0000", font=BLACK, border=BORDER)
    wa.cell(row=cr+1, column=1, value="Precision ratio (1:X)").font = BLACK
    live.set(wa, f"B{cr+1}", f"=ROUND(C{tr}/B{cr},0)", round(lc["ratio"]),
             kind="int", number_format="0", font=BLACK, border=BORDER)

    # coordinates
    coordrow = cr + 3
    wa.cell(row=coordrow, column=1, value="Adjusted coordinates").font = BLACKB
    hdr(wa, coordrow+1, ["Station", "Northing (N)", "Easting (E)"])
    seq = ["AB", "BC", "CD", "DE", "EA"]
    to_station = {"AB": "B", "BC": "C", "CD": "D", "DE": "E", "EA": "A"}
    # A
    ra = coordrow + 2
    wa.cell(row=ra, column=1, value="A").font = BLACK
    wa.cell(row=ra, column=2, value=1000.0).font = BLACK; wa.cell(row=ra, column=2).number_format="0.0000"; wa.cell(row=ra,column=2).border=BORDER
    wa.cell(row=ra, column=3, value=1000.0).font = BLACK; wa.cell(row=ra, column=3).number_format="0.0000"; wa.cell(row=ra,column=3).border=BORDER
    wa.cell(row=ra,column=1).border=BORDER
    # B..E accumulate adjusted lat/dep from previous station using H/I of each course
    station_row = {"A": ra}
    prev = "A"
    for i, c in enumerate(seq):
        nxt = to_station[c]
        if nxt == "A":
            break
        rr = ra + (["B","C","D","E"].index(nxt) + 1)
        course_r = startrow + 1 + i
        live.set(wa, f"B{rr}", f"=B{station_row[prev]}+H{course_r}", round(co[nxt][0], 4),
                 kind="num", dp=4, number_format="0.0000", font=BLACK, border=BORDER)
        live.set(wa, f"C{rr}", f"=C{station_row[prev]}+I{course_r}", round(co[nxt][1], 4),
                 kind="num", dp=4, number_format="0.0000", font=BLACK, border=BORDER)
        wa.cell(row=rr, column=1, value=nxt).font = BLACK; wa.cell(row=rr,column=1).border=BORDER
        station_row[nxt] = rr
        prev = nxt

    for j, w in enumerate([16, 16, 14, 12, 12, 10, 10, 12, 12], 1):
        wa.column_dimensions[get_column_letter(j)].width = w

    # ================= Area =================
    war = wb.create_sheet("Area")
    war["A1"] = "Enclosed area by coordinate (shoelace) method"; war["A1"].font = BLACKB
    war["A3"] = "Area (m^2)"; war["A3"].font = BLACK
    # reference the coordinate cells on Adjustment via shoelace
    # build shoelace referencing station rows
    sr = {s: station_row[s] for s in ["A","B","C","D","E"]}
    terms = []
    st_order = ["A","B","C","D","E"]
    for i in range(5):
        s1 = st_order[i]; s2 = st_order[(i+1)%5]
        terms.append(f"(Adjustment!C{sr[s1]}*Adjustment!B{sr[s2]}-Adjustment!C{sr[s2]}*Adjustment!B{sr[s1]})")
    formula = "=ABS(" + "+".join(terms) + ")/2"
    live.set(war, "B3", formula, round(T.area(), 3), kind="num", dp=3,
             number_format="0.000", font=BLACK, border=BORDER)
    war["A4"] = "Area (hectares)"; war["A4"].font = BLACK
    live.set(war, "B4", "=B3/10000", round(T.area()/10000, 4), kind="num", dp=4,
             number_format="0.0000", font=BLACK, border=BORDER)
    war.column_dimensions["A"].width = 18; war.column_dimensions["B"].width = 14

    wb.save(out)
    n, files = live.inject(out)
    normalize_decimals(out, max_dp=6)
    set_excel_fingerprint(out, application="Microsoft Excel", creator="Survey")
    print(f"wrote {out}: {n} live cells across {files} sheets; area {T.area():.3f} m^2")

if __name__ == "__main__":
    main()

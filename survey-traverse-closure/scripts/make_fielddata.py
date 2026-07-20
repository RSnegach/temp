# -*- coding: utf-8 -*-
"""
Build the keyed field-data input workbook (traverse_field_data.xlsx): the raw
crew values exactly as booked, in their original conventions (bearing, deflection
angle, slope distance). This mirrors the scanned field-book page for analysts who
prefer to pull from a sheet. It does NOT pre-reduce anything.
Black text, neutral header, no reduced/answer values.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from xlsx_live import normalize_decimals, set_excel_fingerprint

thin = Side(style="thin", color="000000")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HDR = PatternFill("solid", fgColor="D9D9D9")
BLACK = Font(color="000000"); BLACKB = Font(color="000000", bold=True)
CEN = Alignment(horizontal="center"); LEF = Alignment(horizontal="left")

def hdr(ws, row, labels):
    for j, t in enumerate(labels, 1):
        c = ws.cell(row=row, column=j, value=t)
        c.font = BLACKB; c.fill = HDR; c.border = BORDER; c.alignment = CEN

def main(out="traverse_field_data.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Field Notes"
    ws["A1"] = "Parcel BLA-7  |  Closed loop A-B-C-D-E-A  |  raw field book (as booked)"
    ws["A1"].font = BLACKB

    hdr(ws, 3, ["Station", "Angle as booked", "Convention"])
    angs = [
        ("A", "98-00-18", "interior angle"),
        ("B", "104-16-21", "interior angle"),
        ("C", "71-01-56 R", "DEFLECTION angle, turned right"),
        ("D", "108-51-30", "interior angle"),
        ("E", "119-54-13", "interior angle"),
    ]
    r = 4
    for st, a, conv in angs:
        ws.cell(row=r,column=1,value=st).font=BLACK
        ws.cell(row=r,column=2,value=a).font=BLACK
        ws.cell(row=r,column=3,value=conv).font=BLACK
        for c in range(1,4): ws.cell(row=r,column=c).border=BORDER
        ws.cell(row=r,column=3).alignment=LEF
        r += 1

    ws["A10"] = "Course distances as booked"; ws["A10"].font = BLACKB
    hdr(ws, 11, ["Course", "Distance as booked", "Convention"])
    dists = [
        ("AB", "199.75 m", "horizontal (EDM)"),
        ("BC", "176.55 m", "horizontal"),
        ("CD", "195.37 m", "SLOPE distance, vertical angle 5-00-00 above horizontal"),
        ("DE", "136.50 m", "horizontal"),
        ("EA", "195.97 m", "horizontal"),
    ]
    r = 12
    for c_, d, conv in dists:
        ws.cell(row=r,column=1,value=c_).font=BLACK
        ws.cell(row=r,column=2,value=d).font=BLACK
        ws.cell(row=r,column=3,value=conv).font=BLACK
        for c in range(1,4): ws.cell(row=r,column=c).border=BORDER
        ws.cell(row=r,column=3).alignment=LEF
        r += 1

    ws["A19"] = "Control and start"; ws["A19"].font = BLACKB
    ws["A20"] = "AB direction (from monument MON-3)"; ws["A20"].font = BLACK
    ws["B20"] = "bearing N 64-07-45 E"; ws["B20"].font = BLACK
    ws["A21"] = "Start coordinates at A"; ws["A21"].font = BLACK
    ws["B21"] = "N 1000.000, E 1000.000"; ws["B21"].font = BLACK

    for j,w in enumerate([10,22,48],1): ws.column_dimensions[get_column_letter(j)].width=w
    wb.save(out)
    normalize_decimals(out, max_dp=6)
    set_excel_fingerprint(out, application="Microsoft Excel", creator="Survey")
    print("wrote", out)

if __name__ == "__main__":
    main()

# -*- coding: utf-8 -*-
"""Input file 2: Decision Rules workbook (.xlsx). Ordered rule engine with explicit
operators + precedence. Holds decision logic only (no validation ranges, no cases)."""
import suite as S
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY="1F385C"; WHITE="FFFFFF"; LT="EEF2F7"; AMBER="FBE9D0"
thin=Side(style="thin",color="C3CAD6"); B=Border(left=thin,right=thin,top=thin,bottom=thin)
HF=Font(name="Calibri",size=10,bold=True,color=WHITE); HFill=PatternFill("solid",fgColor=NAVY)
CF=Font(name="Calibri",size=10,color="1A1A1A"); TF=Font(name="Calibri",size=14,bold=True,color=NAVY)
SF=Font(name="Calibri",size=9,italic=True,color="5A6472"); BF=Font(name="Calibri",size=10,bold=True,color="1A1A1A")
Z=PatternFill("solid",fgColor=LT); AM=PatternFill("solid",fgColor=AMBER)

wb=Workbook()

def hdr(ws,row,headers,start=1):
    for i,h in enumerate(headers):
        c=ws.cell(row=row,column=start+i,value=h); c.font=HF; c.fill=HFill; c.border=B
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
def wtab(ws,top,headers,rows,widths,center=None):
    center=center or []
    hdr(ws,top,headers)
    for ri,row in enumerate(rows):
        for ci,val in enumerate(row):
            c=ws.cell(row=top+1+ri,column=1+ci,value=val); c.font=CF; c.border=B
            c.alignment=Alignment(horizontal="center" if ci in center else "left",vertical="center",wrap_text=True)
            if ri%2==1: c.fill=Z
    for i,w in enumerate(widths): ws.column_dimensions[get_column_letter(1+i)].width=w
    return top+1+len(rows)

# ---- Sheet 1: Evaluation_Order (the ordered rule table) ----
ws=wb.active; ws.title="Rule_Set"
ws["A1"]="LoanPreQual Decision Rule Set"; ws["A1"].font=TF
ws["A2"]="LPQ-482-RULES  |  Release 2.4. Rules are evaluated top to bottom. The FIRST rule whose condition is true returns its outcome; no later rule is evaluated. Applies only to records that passed field validation."
ws["A2"].font=SF; ws["A2"].alignment=Alignment(wrap_text=True); ws.row_dimensions[2].height=42
ws.merge_cells("A2:F2")
rows=[
 ["R1","1","employment_status is Unemployed","DECLINE","no qualifying income source","Applies regardless of all other fields."],
 ["R2","2","credit_score < 620","DECLINE","credit score below minimum","Strict less-than. 620 does not decline here."],
 ["R3","3","existing_debt_ratio > 43.0","DECLINE","debt-to-income above maximum","Strict greater-than. 43.0 exactly does not decline."],
 ["R4","4","loan_amount > 5 x annual_income","DECLINE","loan exceeds income multiple","Strict greater-than. Loan equal to 5x income passes. Multiple is 5."],
 ["R5","5","credit_score >= 720 AND existing_debt_ratio <= 36.0","APPROVE_PRIME","prime tier","Both conditions inclusive. Both must hold."],
 ["R6","6","credit_score >= 660","APPROVE_STANDARD","standard tier","Inclusive. Reached only if not prime."],
 ["R7","7","(no condition, default)","REFER","manual underwriting review","Catch-all for valid records not matched above."],
]
last=wtab(ws,4,["Rule","Order","Condition","Outcome","Reason returned","Operator notes"],rows,
          [7,7,34,18,26,40],center=[0,1])
ws.freeze_panes="A5"

# ---- Sheet 2: Thresholds (named constants, so BVA/DB have explicit anchors) ----
ws2=wb.create_sheet("Thresholds")
ws2["A1"]="Decision Thresholds"; ws2["A1"].font=TF
ws2["A2"]="Named business thresholds referenced by the rule set. These are business limits, distinct from field validation ranges (see LPQ-482-VAL)."
ws2["A2"].font=SF; ws2.merge_cells("A2:D2"); ws2.row_dimensions[2].height=30
thr=[
 ["Minimum lending score","620","credit_score","R2 uses strict < ; at 620 the applicant is not declined by R2"],
 ["Maximum debt-to-income","43.0","existing_debt_ratio (%)","R3 uses strict > ; 43.0 exactly is allowed"],
 ["Loan-to-income multiple","5","x annual_income","R4 uses strict > ; loan = 5x income is allowed"],
 ["Prime score floor","720","credit_score","R5 uses >= (inclusive)"],
 ["Prime DTI ceiling","36.0","existing_debt_ratio (%)","R5 uses <= (inclusive)"],
 ["Standard score floor","660","credit_score","R6 uses >= (inclusive)"],
]
wtab(ws2,4,["Threshold","Value","Applies to","Operator / boundary note"],thr,[24,10,22,46],center=[1])
ws2.freeze_panes="A5"

# ---- Sheet 3: Precedence_Examples (worked precedence, NOT test cases) ----
ws3=wb.create_sheet("Precedence_Notes")
ws3["A1"]="Precedence Notes"; ws3["A1"].font=TF
ws3["A2"]="How overlapping conditions resolve. These illustrate ordering only; they are not a test set and do not enumerate inputs."
ws3["A2"].font=SF; ws3.merge_cells("A2:C2"); ws3.row_dimensions[2].height=30
prec=[
 ["Unemployed applicant who also has a low score","R1 returns first","DECLINE (no qualifying income source), not the score reason"],
 ["Low score AND high debt-to-income","R2 is earlier than R3","DECLINE (credit score below minimum)"],
 ["Score 700, debt-to-income 30 percent","Not prime (needs >=720), clears standard floor","APPROVE_STANDARD"],
 ["Score 640, debt-to-income within limit","Above 620 floor, below 660 standard floor","REFER"],
 ["Strong score but debt-to-income 37 percent","Fails prime DTI ceiling of 36.0, clears standard","APPROVE_STANDARD not APPROVE_PRIME"],
]
wtab(ws3,4,["Situation","Why","Returned outcome"],prec,[34,28,34])
ws3.freeze_panes="A5"

wb.properties.creator="Product Management"; wb.properties.title="LoanPreQual Decision Rules"; wb.properties.lastModifiedBy="Product Management"
out="inputs/LPQ-482-RULES_Decision_Rules.xlsx"; wb.save(out); print("saved",out)

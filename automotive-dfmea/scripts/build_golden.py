# -*- coding: utf-8 -*-
"""Golden DFMEA workbook with LIVE formulas.
Sheets:
  DFMEA         - one row per active failure mode; D via live VLOOKUP of the control,
                  AP via live lookup of the AP table, action via live IF, ranked.
  Detection_Map - control -> D reference (VLOOKUP target).
  AP_Table      - expanded S/O/D-band lookup (INDEX/MATCH target for AP).
Headers black, one accent sheet (AP_Table), no italic subtitle rows, no navy.
Caches injected, tails stripped, fingerprint set.
"""
import dfmea as D
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from xlsx_live import LiveCells, normalize_decimals, set_excel_fingerprint

BLACK="000000"; WHITE="FFFFFF"; ACCENT="2A4A6B"; LT="EEF2F6"
HFILL_H="F5CBC6"; HFILL_M="FBE8C8"; HFILL_L="C9E6D4"
thin=Side(style="thin",color="C3CDD6"); B=Border(left=thin,right=thin,top=thin,bottom=thin)
def F(sz=10,b=False,c="1A1A1A"): return Font(name="Calibri",size=sz,bold=b,color=c)
def MONO(sz=9): return Font(name="Consolas",size=sz,color="1A1A1A")

live=LiveCells()
rows=D.score()
wb=Workbook()

def hdr(ws,row,headers,fill=BLACK,start=1):
    for i,h in enumerate(headers):
        c=ws.cell(row,start+i,h); c.fill=PatternFill("solid",fgColor=fill)
        c.font=Font(name="Calibri",size=10,bold=True,color=WHITE)
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=B

# ---------- Detection_Map sheet ----------
dm=wb.active; dm.title="Detection_Map"
dm["A1"]="Detection Rating Reference (control -> D)"; dm["A1"].font=F(12,True,"333333")
hdr(dm,2,["current_detection_control","D"])
dmap=sorted(D.DETECTION_MAP.items(), key=lambda kv:kv[1])
for i,(k,v) in enumerate(dmap):
    r=3+i
    dm.cell(r,1,k).font=MONO(); dm.cell(r,2,v).font=MONO()
    dm.cell(r,1).border=B; dm.cell(r,2).border=B; dm.cell(r,2).alignment=Alignment(horizontal="center")
dm_first,dm_last=3,2+len(dmap)
dm.column_dimensions["A"].width=40; dm.column_dimensions["B"].width=6
dm.freeze_panes="A3"
DMAP_RNG=f"Detection_Map!$A${dm_first}:$B${dm_last}"

# ---------- AP_Table sheet (expanded to band-edge rows for INDEX/MATCH) ----------
apt=wb.create_sheet("AP_Table")
apt["A1"]="Action Priority Lookup (AIAG-VDA)"; apt["A1"].font=F(12,True,"333333")
hdr(apt,2,["S_lo","S_hi","O_lo","O_hi","D_lo","D_hi","AP"],fill=ACCENT)
for i,(slo,shi,olo,ohi,dlo,dhi,ap) in enumerate(D.AP_ROWS):
    r=3+i
    for ci,v in enumerate([slo,shi,olo,ohi,dlo,dhi,ap]):
        c=apt.cell(r,1+ci,v); c.font=MONO(); c.border=B; c.alignment=Alignment(horizontal="center")
apt_first,apt_last=3,2+len(D.AP_ROWS)
for col in "ABCDEFG": apt.column_dimensions[col].width=6
apt.column_dimensions["G"].width=5
apt.freeze_panes="A3"

# ---------- DFMEA main sheet ----------
ws=wb.create_sheet("DFMEA")
ws["A1"]="Fuel Delivery System - Design FMEA (AIAG-VDA)"; ws["A1"].font=F(14,True,ACCENT)
cols=["Rank","fm_id","Item","Function","Failure Mode","Effect","S","O","Detection Control","D","AP","Action"]
hdr(ws,2,cols)
# column letters: A Rank B id C item D func E mode F effect G S H O I control J D K AP L action
r0=3
for i,r in enumerate(rows):
    rr=r0+i
    ws.cell(rr,1,r["rank"]).font=MONO();
    ws.cell(rr,2,r["id"]).font=F(9,True)
    ws.cell(rr,3,r["item"]).font=F(9); ws.cell(rr,4,r["function"]).font=F(9)
    ws.cell(rr,5,r["mode"]).font=F(9); ws.cell(rr,6,r["effect"]).font=F(9)
    ws.cell(rr,7,r["S"]).font=MONO(); ws.cell(rr,8,r["O"]).font=MONO()
    ws.cell(rr,9,r["control"]).font=F(9)
    # D = live VLOOKUP of control against Detection_Map
    live.set(ws,f"J{rr}",f'=VLOOKUP(I{rr},{DMAP_RNG},2,FALSE)', r["D"], kind="num", dp=0,
             number_format="0", font=MONO(), align=Alignment(horizontal="center"))
    # AP = live lookup of the AP_Table row whose S/O/D bands contain this row's
    # values. SUMPRODUCT picks the matching row index (bands are disjoint, so
    # exactly one matches); INDEX returns its AP letter. SUMPRODUCT evaluates as
    # an ordinary formula (no Ctrl+Shift+Enter), so it works in every Excel.
    A=f'AP_Table!$A${apt_first}:$A${apt_last}'; Bc=f'AP_Table!$B${apt_first}:$B${apt_last}'
    C=f'AP_Table!$C${apt_first}:$C${apt_last}'; Dc=f'AP_Table!$D${apt_first}:$D${apt_last}'
    E=f'AP_Table!$E${apt_first}:$E${apt_last}'; Fc=f'AP_Table!$F${apt_first}:$F${apt_last}'
    Gc=f'AP_Table!$G${apt_first}:$G${apt_last}'
    ap_formula=(f'=INDEX({Gc},SUMPRODUCT((ROW({A})-{apt_first-1})*'
                f'({A}<=G{rr})*({Bc}>=G{rr})*({C}<=H{rr})*({Dc}>=H{rr})*({E}<=J{rr})*({Fc}>=J{rr})))')
    live.set(ws,f"K{rr}",ap_formula, r["ap"], kind="str", font=F(9,True),
             align=Alignment(horizontal="center"))
    # Action = live IF on AP
    live.set(ws,f"L{rr}",'=IF(K{0}="H","Yes",IF(K{0}="M","Review","No"))'.format(rr), r["action"],
             kind="str", font=F(9), align=Alignment(horizontal="center"))
    # color AP cell by value
    fillc={"H":HFILL_H,"M":HFILL_M,"L":HFILL_L}[r["ap"]]
    ws.cell(rr,11).fill=PatternFill("solid",fgColor=fillc)
    for c in range(1,13):
        ws.cell(rr,c).border=B
        if ws.cell(rr,c).alignment is None or ws.cell(rr,c).alignment.horizontal is None:
            ws.cell(rr,c).alignment=Alignment(vertical="center",wrap_text=True)
widths=[5,7,15,20,26,26,4,4,26,4,5,8]
for i,w in enumerate(widths): ws.column_dimensions[get_column_letter(1+i)].width=w
ws.freeze_panes="C3"

# order sheets: DFMEA first
wb.move_sheet("DFMEA", -(len(wb.sheetnames)-1))
wb.properties.creator="Product Engineering"; wb.properties.title="Fuel Delivery System DFMEA"; wb.properties.lastModifiedBy="Product Engineering"
OUT="golden/FDS-DFMEA_Fuel_Delivery.xlsx"; wb.save(OUT)

# inject caches + clean
nc,nf=live.inject(OUT)
normalize_decimals(OUT); set_excel_fingerprint(OUT, creator="Product Engineering")
from collections import Counter
print("saved",OUT,f"| live cells {nc} across {nf} sheets")
print("AP dist:", dict(Counter(r['ap'] for r in rows)), "| action Yes:", sum(1 for r in rows if r['action']=='Yes'))

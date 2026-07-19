# -*- coding: utf-8 -*-
"""Golden deliverable: LPQ-482_Test_Suite.xlsx built from suite.py.
Sheets: Test_Cases (all 55 with full input vector + expected result + coverage tag),
Coverage_Summary (counts by technique + traceability to rules/fields), Legend."""
import suite as S
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter

NAVY="1F385C"; WHITE="FFFFFF"; LT="EEF2F7"
TECHFILL={"EP":"E8F0FE","BVA":"FDECEC","DT":"E9F6EC","DB":"FBF3E0","PATH":"ECE6F5"}
CODEFILL={"APPROVE_PRIME":"CDE8D2","APPROVE_STANDARD":"E4F1E7","REFER":"FBF0D5","DECLINE":"F7D9D5","REJECT":"E2E4E8"}
thin=Side(style="thin",color="C3CAD6"); B=Border(left=thin,right=thin,top=thin,bottom=thin)
HF=Font(name="Calibri",size=10,bold=True,color=WHITE); HFill=PatternFill("solid",fgColor=NAVY)
CF=Font(name="Calibri",size=10,color="1A1A1A"); MONO=Font(name="Consolas",size=9,color="1A1A1A")
TF=Font(name="Calibri",size=14,bold=True,color=NAVY); SF=Font(name="Calibri",size=9,italic=True,color="5A6472")
BF=Font(name="Calibri",size=10,bold=True,color="1A1A1A")

cases=S.build_all()
wb=Workbook()

def hdr(ws,row,headers):
    for i,h in enumerate(headers):
        c=ws.cell(row=row,column=1+i,value=h); c.font=HF; c.fill=HFill; c.border=B
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)

# ---------------- Test_Cases ----------------
ws=wb.active; ws.title="Test_Cases"
ws["A1"]="LoanPreQual Decision Service - Test Suite"; ws["A1"].font=TF
ws["A2"]=("Derived from LPQ-482 Feature Specification, LPQ-482-RULES Decision Rules, and LPQ-482-VAL Field "
          "Validation Reference. Techniques: EP equivalence partitioning, BVA boundary value analysis, "
          "DT decision table (one per rule), DB decision-boundary analysis. Every case holds non-target fields "
          "at the baseline applicant.")
ws["A2"].font=SF; ws.merge_cells("A2:N2"); ws.row_dimensions[2].height=40
FIELDS=S.FIELD_ORDER
headers=["Case ID","Technique","Target / rationale"]+FIELDS+["Expected result"]
ws["A2"].value=(ws["A2"].value+" PATH = basis-path coverage of the decision-flow graph.")
hdr(ws,4,headers)
r=5
for c in cases:
    ws.cell(r,1,c["id"]).font=BF
    tcell=ws.cell(r,2,c["technique"]); tcell.font=CF; tcell.fill=PatternFill("solid",fgColor=TECHFILL[c["technique"]])
    ws.cell(r,3,c["target"]).font=CF
    for j,f in enumerate(FIELDS):
        v=c["vector"][f]
        cell=ws.cell(r,4+j,v); cell.font=MONO; cell.alignment=Alignment(horizontal="center")
        if f=="existing_debt_ratio": cell.number_format="0.0"
    code=c["expected"].split(":")[0]
    ecell=ws.cell(r,4+len(FIELDS),c["expected"]); ecell.font=CF; ecell.fill=PatternFill("solid",fgColor=CODEFILL.get(code,"FFFFFF"))
    for col in range(1,5+len(FIELDS)):
        cc=ws.cell(r,col); cc.border=B
        if cc.alignment is None or cc.alignment.horizontal is None:
            cc.alignment=Alignment(vertical="center")
    r+=1
widths=[9,10,40, 11,13,12,12,17,16,15, 40]
for i,w in enumerate(widths): ws.column_dimensions[get_column_letter(1+i)].width=w
ws.freeze_panes="D5"

# ---------------- Coverage_Summary ----------------
ws2=wb.create_sheet("Coverage_Summary")
ws2["A1"]="Coverage Summary"; ws2["A1"].font=TF
ws2["A2"]="Case counts by technique and outcome, plus traceability that every field range limit and every decision rule is exercised."
ws2["A2"].font=SF; ws2.merge_cells("A2:D2"); ws2.row_dimensions[2].height=28

by_tech=Counter(c["technique"] for c in cases)
hdr(ws2,4,["Technique","Cases","What it covers"])
tech_desc={"EP":"One representative per equivalence class (valid + invalid), per field",
           "BVA":"min-1 / min / max / max+1 for each of 5 bounded numeric fields",
           "DT":"One first-match case per decision rule R1..R7",
           "DB":"Each decision threshold at and across the boundary"}
tech_desc["PATH"]="One case per independent basis path of the decision-flow graph, V(G)=8"
rr=5
for t in ["EP","BVA","DT","DB","PATH"]:
    ws2.cell(rr,1,t).font=BF; ws2.cell(rr,2,by_tech[t]).font=CF; ws2.cell(rr,3,tech_desc[t]).font=CF
    for col in range(1,4): ws2.cell(rr,col).border=B
    rr+=1
ws2.cell(rr,1,"TOTAL").font=BF; ws2.cell(rr,2,len(cases)).font=BF
for col in range(1,4): ws2.cell(rr,col).border=B; ws2.cell(rr,col).fill=PatternFill("solid",fgColor=LT)
rr+=2

by_code=Counter(c["expected"].split(":")[0] for c in cases)
hdr(ws2,rr,["Outcome","Cases"]); rr+=1
for code in ["APPROVE_PRIME","APPROVE_STANDARD","REFER","DECLINE","REJECT"]:
    ws2.cell(rr,1,code).font=CF; ws2.cell(rr,1).fill=PatternFill("solid",fgColor=CODEFILL[code])
    ws2.cell(rr,2,by_code[code]).font=CF
    for col in range(1,3): ws2.cell(rr,col).border=B
    rr+=1
for w,c in zip([26,10,60],"ABC"): ws2.column_dimensions[c].width=w

# ---------------- Rule_Traceability ----------------
ws3=wb.create_sheet("Rule_Traceability")
ws3["A1"]="Decision Rule Traceability"; ws3["A1"].font=TF
ws3["A2"]="Each decision rule mapped to the cases that exercise it (decision-table plus boundary cases)."
ws3["A2"].font=SF; ws3.merge_cells("A2:C2")
hdr(ws3,4,["Rule","Cases exercising it","Expected outcome"])
rule_map={
 "R1":("DT-001","DECLINE no qualifying income source"),
 "R2":("DT-002, DB-001, DB-002","DECLINE below min score (619) / 620 falls through"),
 "R3":("DT-003, DB-009, DB-010","DECLINE DTI above max (43.1) / 43.0 allowed"),
 "R4":("DT-004, DB-011, DB-012","DECLINE loan over 5x income / exactly 5x allowed"),
 "R5":("DT-005, DB-005, DB-006, DB-007, DB-008, BVA-011","APPROVE_PRIME at score>=720 and DTI<=36.0"),
 "R6":("DT-006, DB-003, DB-004","APPROVE_STANDARD at score>=660 not prime"),
 "R7":("DT-007, DB-002, DB-003","REFER when 620<=score<660"),
}
rr=5
for rule,(cs,out) in rule_map.items():
    ws3.cell(rr,1,rule).font=BF; ws3.cell(rr,2,cs).font=CF; ws3.cell(rr,3,out).font=CF
    for col in range(1,4): ws3.cell(rr,col).border=B
    rr+=1
for w,c in zip([8,42,46],"ABC"): ws3.column_dimensions[c].width=w

# ---------------- Branch_Coverage ----------------
ws4=wb.create_sheet("Branch_Coverage")
ws4["A1"]="Decision-Flow Branch Coverage"; ws4["A1"].font=TF
vg,N,E=S.cfg_edge_node_check()
ws4["A2"]=(f"Decision-flow graph: {len(S.DECISION_NODES)} decision nodes, {2*len(S.DECISION_NODES)} branches, "
           f"cyclomatic complexity V(G) = {S.CYCLOMATIC} (cross-checked E-N+2 = {vg}). "
           "Every branch below is exercised by at least one case; R5 is decomposed into D5 (score gate) and D6 (DTI gate).")
ws4["A2"].font=SF; ws4.merge_cells("A2:D2"); ws4.row_dimensions[2].height=42
node_q={"D1":"employment = Unemployed","D2":"credit_score < 620","D3":"existing_debt_ratio > 43.0",
        "D4":"loan_amount > 5x income","D5":"credit_score >= 720 (R5 gate 1)",
        "D6":"existing_debt_ratio <= 36.0 (R5 gate 2)","D7":"credit_score >= 660"}
cov=S.branch_coverage()
hdr(ws4,4,["Decision node","Branch","Covering case(s)","Leads to"])
lead={("D1",True):"DECLINE R1",("D1",False):"-> D2",("D2",True):"DECLINE R2",("D2",False):"-> D3",
      ("D3",True):"DECLINE R3",("D3",False):"-> D4",("D4",True):"DECLINE R4",("D4",False):"-> D5",
      ("D5",True):"-> D6",("D5",False):"-> D7 (not prime)",("D6",True):"APPROVE_PRIME R5",("D6",False):"-> D7",
      ("D7",True):"APPROVE_STANDARD R6",("D7",False):"REFER R7"}
rr=5
for d in S.DECISION_NODES:
    for b in (True,False):
        cases_hit=cov.get((d,b),[])
        # show a compact sample (first 3) so the cell stays readable
        shown=", ".join(cases_hit[:3]) + (f"  (+{len(cases_hit)-3} more)" if len(cases_hit)>3 else "")
        ws4.cell(rr,1,f"{d}: {node_q[d]}").font=CF
        ws4.cell(rr,2,"TRUE" if b else "FALSE").font=CF
        ws4.cell(rr,3,shown or "NONE").font=CF
        ws4.cell(rr,4,lead[(d,b)]).font=CF
        for col in range(1,5):
            ws4.cell(rr,col).border=B
            if not cases_hit: ws4.cell(rr,col).fill=PatternFill("solid",fgColor="F7D9D5")
        rr+=1
for w,c in zip([34,10,34,26],"ABCD"): ws4.column_dimensions[c].width=w
ws4.freeze_panes="A5"

wb.properties.creator="Quality Assurance"; wb.properties.title="LoanPreQual Test Suite"; wb.properties.lastModifiedBy="Quality Assurance"
out="golden/LPQ-482_Test_Suite.xlsx"; wb.save(out); print("saved",out,"with",len(cases),"cases")

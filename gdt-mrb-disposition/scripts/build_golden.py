# -*- coding: utf-8 -*-
"""
Build the golden disposition deliverable: BRKT-4471_Inspection_Disposition.xlsx.

Every DERIVED column (bonus, datum shift, total allowed, actual deviation,
virtual condition, size check, disposition) is a LIVE Excel formula referencing
the transcribed given cells and the Datums tab, so the sheet recalculates if a
measurement changes. The transcribed given values (from the drawing callouts and
the CMM report) are static inputs. Cached results are injected so the numbers
show without opening Excel; decimals normalized; fingerprint scrubbed.

Formatting per Geranium rules: black text everywhere, one neutral light-gray
header theme across all sheets, no colored header fills, no italic subtitle rows.
Disposition column carries a semantic conditional-format (accept/mrb/reject),
which is allowed because it encodes data.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
import gdt
from xlsx_live import LiveCells, normalize_decimals, set_excel_fingerprint

thin = Side(style="thin", color="000000")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HDRFILL = PatternFill("solid", fgColor="D9D9D9")
BLACK = Font(color="000000")
BLACKB = Font(color="000000", bold=True)
CEN = Alignment(horizontal="center", vertical="center")
LEF = Alignment(horizontal="left", vertical="center")
# semantic disposition fills (data, not decoration)
F_OK = PatternFill("solid", fgColor="C6EFCE")   # green
F_MRB = PatternFill("solid", fgColor="FFEB9C")  # amber
F_REJ = PatternFill("solid", fgColor="FFC7CE")  # red

def hdr(ws, row, labels, start=1):
    for j, t in enumerate(labels, start):
        c = ws.cell(row=row, column=j, value=t)
        c.font = BLACKB; c.fill = HDRFILL; c.border = BORDER; c.alignment = CEN

def main(out="BRKT-4471_Inspection_Disposition.xlsx"):
    rows = gdt.evaluate()
    # index golden results by id for cached injection
    G = {r["id"]: r for r in rows}
    live = LiveCells()

    wb = openpyxl.Workbook()

    # ================= sheet: Datums =================
    wd = wb.active
    wd.title = "Datums"
    wd["A1"] = "Datum Features of Size"; wd["A1"].font = BLACKB
    hdr(wd, 3, ["Datum", "Type", "MMC (mm)", "Measured (mm)", "Departure from MMC (mm)"])
    drows = [("B", "internal", gdt.DATUMS["B"]["mmc"], gdt.DATUMS["B"]["actual"]),
             ("C", "internal", gdt.DATUMS["C"]["mmc"], gdt.DATUMS["C"]["actual"])]
    for i, (name, typ, mmc, meas) in enumerate(drows):
        r = 4 + i
        for col, val, al in [(1, name, CEN), (2, typ, CEN), (3, mmc, CEN), (4, meas, CEN)]:
            c = wd.cell(row=r, column=col, value=val); c.font = BLACK; c.border = BORDER; c.alignment = al
            if col in (3, 4): c.number_format = "0.00"
        # departure = internal ? meas-mmc : mmc-meas  (live formula)
        dep = round(gdt.datum_shift(name), 4)
        live.set(wd, f"E{r}", f'=IF(B{r}="internal",D{r}-C{r},C{r}-D{r})', dep,
                 kind="num", dp=3, number_format="0.000", font=BLACK, border=BORDER, align=CEN)
    for j, w in enumerate([9, 12, 12, 14, 24], 1):
        wd.column_dimensions[get_column_letter(j)].width = w

    # ================= sheet: Disposition =================
    ws = wb.create_sheet("Disposition")
    ws["A1"] = "BRKT-4471 Mounting Bracket  |  Inspection Disposition  |  Serial 4471-0007"
    ws["A1"].font = BLACKB
    cols = ["Feature", "Type", "Char", "Modifier", "Datum @MMC", "Meas Size",
            "Size MMC", "Size LMC", "Dev X", "Dev Y", "Surf Dev", "Stated Tol",
            "Bonus", "Datum Shift", "Total Allowed", "Actual Dev", "Virtual Cond",
            "Size Check", "Disposition"]
    hdr(ws, 3, cols)
    # column letters
    C = {name: get_column_letter(i) for i, name in enumerate(cols, 1)}

    # order rows same as gdt.FEATURES
    order = [f["id"] for f in gdt.FEATURES]
    feat_by_id = {f["id"]: f for f in gdt.FEATURES}

    r = 4
    disp_first = r
    for fid in order:
        f = feat_by_id[fid]
        g = G[fid]
        is_pos = f["char"] == "position"
        # ---- given (static) ----
        given = {
            "Feature": fid,
            "Type": f["kind"],
            "Char": f["char"],
            "Modifier": (f.get("modifier") or "-") if is_pos else "-",
            "Datum @MMC": (f.get("datum_at_mmc") or "") if is_pos else "",
            "Meas Size": f.get("meas_size", "") if is_pos else "",
            "Size MMC": f.get("size_mmc", "") if is_pos else "",
            "Size LMC": f.get("size_lmc", "") if is_pos else "",
            "Dev X": f.get("dx", "") if is_pos else "",
            "Dev Y": f.get("dy", "") if is_pos else "",
            "Surf Dev": "" if is_pos else f.get("prof_dev", ""),
            "Stated Tol": f["geo_tol"],
        }
        for name, val in given.items():
            c = ws.cell(row=r, column=cols.index(name) + 1, value=val)
            c.font = BLACK; c.border = BORDER
            c.alignment = LEF if name in ("Feature", "Type", "Char") else CEN
            if name in ("Meas Size", "Size MMC", "Size LMC", "Stated Tol") and val != "":
                c.number_format = "0.00"
            if name in ("Dev X", "Dev Y", "Surf Dev") and val != "":
                c.number_format = "0.000"

        # ---- derived (live formulas) ----
        M, N, O, P, Q, R, S = (C["Bonus"], C["Datum Shift"], C["Total Allowed"],
                               C["Actual Dev"], C["Virtual Cond"], C["Size Check"],
                               C["Disposition"])
        cM, cF, cG, cH, cI, cJ, cK, cL, cCh, cB, cD, cE = (
            C["Char"], C["Meas Size"], C["Size MMC"], C["Size LMC"], C["Dev X"],
            C["Dev Y"], C["Surf Dev"], C["Stated Tol"], C["Char"], C["Type"],
            C["Modifier"], C["Datum @MMC"])

        # Bonus
        live.set(ws, f"{M}{r}",
                 f'=IF({cCh}{r}<>"position",0,IF({cD}{r}<>"MMC",0,'
                 f'IF({cB}{r}="pin",{cG}{r}-{cF}{r},{cF}{r}-{cG}{r})))',
                 g["bonus"], kind="num", dp=3, number_format="0.000",
                 font=BLACK, border=BORDER, align=CEN)
        # Datum shift (VLOOKUP into Datums departure column)
        live.set(ws, f"{N}{r}",
                 f'=IF({cE}{r}="",0,VLOOKUP({cE}{r},Datums!$A$4:$E$5,5,FALSE))',
                 g["shift"], kind="num", dp=3, number_format="0.000",
                 font=BLACK, border=BORDER, align=CEN)
        # Total allowed
        live.set(ws, f"{O}{r}",
                 f'=IF({cCh}{r}="position",{cL}{r}+{M}{r}+{N}{r},{cL}{r})',
                 g["total"], kind="num", dp=3, number_format="0.000",
                 font=BLACK, border=BORDER, align=CEN)
        # Actual deviation
        live.set(ws, f"{P}{r}",
                 f'=IF({cCh}{r}="position",2*SQRT({cI}{r}^2+{cJ}{r}^2),{cK}{r})',
                 g["actual"], kind="num", dp=3, number_format="0.000",
                 font=BLACK, border=BORDER, align=CEN)
        # Virtual condition
        vc_cached = g["vc"] if g["vc"] is not None else ""
        if is_pos:
            live.set(ws, f"{Q}{r}",
                     f'=IF({cCh}{r}<>"position","",IF({cB}{r}="pin",{cG}{r}+{cL}{r},{cG}{r}-{cL}{r}))',
                     vc_cached, kind="num", dp=3, number_format="0.000",
                     font=BLACK, border=BORDER, align=CEN)
        else:
            c = ws.cell(row=r, column=cols.index("Virtual Cond") + 1, value="")
            c.font = BLACK; c.border = BORDER; c.alignment = CEN
        # Size check
        size_cached = ("n/a" if not is_pos else ("FAIL" if not g["size_ok"] else "PASS"))
        live.set(ws, f"{R}{r}",
                 f'=IF({cCh}{r}<>"position","n/a",'
                 f'IF(AND({cF}{r}>=MIN({cG}{r},{cH}{r}),{cF}{r}<=MAX({cG}{r},{cH}{r})),"PASS","FAIL"))',
                 size_cached, kind="str", font=BLACK, border=BORDER, align=CEN)
        # Disposition
        live.set(ws, f"{S}{r}",
                 f'=IF({R}{r}="FAIL","REJECT (size)",IF({P}{r}<={O}{r},"ACCEPT","MRB"))',
                 g["disp"], kind="str", font=BLACKB, border=BORDER, align=CEN)
        r += 1
    disp_last = r - 1

    widths = [9, 9, 9, 9, 11, 10, 9, 9, 8, 8, 9, 10, 8, 11, 13, 11, 12, 10, 15]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    # semantic conditional formatting on Disposition column
    S = C["Disposition"]
    rng = f"{S}{disp_first}:{S}{disp_last}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="containsText", formula=['"ACCEPT"'], fill=F_OK))
    # containsText for MRB/REJECT via text rules
    from openpyxl.formatting.rule import Rule
    from openpyxl.styles.differential import DifferentialStyle
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"ACCEPT"'], fill=F_OK))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"MRB"'], fill=F_MRB))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"REJECT (size)"'], fill=F_REJ))

    # ================= sheet: Summary =================
    wsu = wb.create_sheet("Summary")
    wsu["A1"] = "Disposition Summary"; wsu["A1"].font = BLACKB
    hdr(wsu, 3, ["Disposition", "Count"])
    from collections import Counter
    counts = Counter(g["disp"] for g in rows)
    labels = ["ACCEPT", "MRB", "REJECT (size)"]
    for i, lab in enumerate(labels):
        rr = 4 + i
        c = wsu.cell(row=rr, column=1, value=lab); c.font = BLACK; c.border = BORDER; c.alignment = LEF
        # live COUNTIF against the Disposition column
        live.set(wsu, f"B{rr}",
                 f'=COUNTIF(Disposition!{S}{disp_first}:{S}{disp_last},A{rr})',
                 counts.get(lab, 0), kind="int", number_format="0",
                 font=BLACK, border=BORDER, align=CEN)
    rr = 4 + len(labels)
    c = wsu.cell(row=rr, column=1, value="Total"); c.font = BLACKB; c.border = BORDER; c.alignment = LEF
    live.set(wsu, f"B{rr}", f"=SUM(B4:B{rr-1})", len(rows), kind="int",
             number_format="0", font=BLACKB, border=BORDER, align=CEN)
    wsu.column_dimensions["A"].width = 16; wsu.column_dimensions["B"].width = 8

    # order sheets: Disposition first, then Datums, then Summary
    wb.move_sheet("Disposition", -(wb.sheetnames.index("Disposition")))
    wb.save(out)
    n, files = live.inject(out)
    normalize_decimals(out, max_dp=6)
    set_excel_fingerprint(out, application="Microsoft Excel", creator="Quality Engineering")
    print(f"wrote {out}: injected {n} live cells across {files} sheet files")
    print("disposition counts:", dict(counts))

if __name__ == "__main__":
    main()

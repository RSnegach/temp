"""
Build the CMM inspection report the gate model receives: measured local size
and measured center offset (dx, dy) from true position for every feature, plus
measured surface deviations for the profile/flatness callouts. Numbers come
straight from gdt.py so the artifact and the golden are consistent.

Black text, plain neutral header, no colored fills, no italic subtitle rows,
per the Geranium formatting rules. This is a raw INPUT artifact (what the
inspector hands over), not the graded deliverable.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import gdt
from xlsx_live import normalize_decimals, set_excel_fingerprint

thin = Side(style="thin", color="000000")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HDRFILL = PatternFill("solid", fgColor="D9D9D9")   # light gray, neutral
BLACK = Font(color="000000")
BLACKB = Font(color="000000", bold=True)

def style_header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = BLACKB
        cell.fill = HDRFILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

def main(out="CMM_inspection_report.xlsx"):
    wb = openpyxl.Workbook()

    # ---------------- sheet 1: feature measurements ----------------
    ws = wb.active
    ws.title = "CMM Results"
    ws["A1"] = "BRKT-4471 Mounting Bracket  |  CMM Inspection Report  |  Serial 4471-0007"
    ws["A1"].font = BLACKB
    hdr = ["Feature", "Type", "Characteristic", "Meas. Local Size (mm)",
           "Center dev X (mm)", "Center dev Y (mm)", "Surface Dev (mm)", "Inspector Note"]
    ws.append([])  # row 2 spacer intentionally blank -> actually we want header row 2
    for j, h in enumerate(hdr, 1):
        ws.cell(row=2, column=j, value=h)
    style_header(ws, 2, len(hdr))

    notes = {
        "H1": "pattern hole", "H2": "pattern hole", "H3": "pattern hole",
        "H4": "pattern hole", "P1": "dowel pin, external feature",
        "R1": "precision bore", "H5": "clearance hole",
        "S1": "profile scan, 240 points", "F1": "flatness, 180 points",
    }
    r = 3
    for f in gdt.FEATURES:
        if f["char"] == "position":
            row = [f["id"], f["kind"], "position", f["meas_size"],
                   f["dx"], f["dy"], "", notes.get(f["id"], "")]
        else:
            row = [f["id"], f["kind"], f["char"], "", "", "",
                   f["prof_dev"], notes.get(f["id"], "")]
        for j, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=j, value=v)
            cell.font = BLACK; cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=8).alignment = Alignment(horizontal="left")
        r += 1

    widths = [10, 10, 14, 20, 16, 16, 15, 26]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w

    # ---------------- sheet 2: datum feature measurements ----------------
    ws2 = wb.create_sheet("Datum Features")
    ws2["A1"] = "Datum Features of Size, Measured Local Size"
    ws2["A1"].font = BLACKB
    hdr2 = ["Datum", "Type", "Size Limits (mm)", "Meas. Local Size (mm)", "Note"]
    for j, h in enumerate(hdr2, 1):
        ws2.cell(row=2, column=j, value=h)
    style_header(ws2, 2, len(hdr2))
    dstat = {
        "B": ("internal (hole)", "10.00 to 10.06", gdt.DATUMS["B"]["actual"], "primary size datum for hole pattern"),
        "C": ("internal (hole)", "6.00 to 6.05", gdt.DATUMS["C"]["actual"], "tertiary datum"),
    }
    r = 3
    for d, (typ, lim, act, note) in dstat.items():
        row = [d, typ, lim, act, note]
        for j, v in enumerate(row, 1):
            cell = ws2.cell(row=r, column=j, value=v)
            cell.font = BLACK; cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")
        ws2.cell(row=r, column=5).alignment = Alignment(horizontal="left")
        r += 1
    for j, w in enumerate([9, 16, 18, 22, 34], 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w

    wb.save(out)
    normalize_decimals(out, max_dp=6)
    set_excel_fingerprint(out, application="Microsoft Excel", creator="Quality Engineering")
    print("wrote", out)

if __name__ == "__main__":
    main()

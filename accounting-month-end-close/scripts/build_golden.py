# -*- coding: utf-8 -*-
"""
Build the golden month-end close workbook: Adjusted_Trial_Balance.xlsx.

Sheets:
  Unadjusted TB : footed balances (values) with the transposition flagged.
  Adjusting Entries : the 7 AJEs plus the transposition restatement note; each
                      AJE's debit=credit checked with a live formula.
  Adjusted TB   : per-account adjusted balance as a LIVE formula = unadjusted
                  (corrected) + sum of AJE effects; totals with live SUM; the
                  balance check debits=credits is a live formula.
  Bank Rec      : adjusted bank vs adjusted book, live.
All cached values come from accounting.py so the golden ties out. Black text,
neutral header, semantic green check on the balanced totals only.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import accounting as A
from xlsx_live import LiveCells, normalize_decimals, set_excel_fingerprint

thin = Side(style="thin", color="000000")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HDR = PatternFill("solid", fgColor="D9D9D9")
BLACK = Font(color="000000"); BLACKB = Font(color="000000", bold=True)
CEN = Alignment(horizontal="center"); LEF = Alignment(horizontal="left")

def hdr(ws, row, labels):
    for j, t in enumerate(labels, 1):
        c = ws.cell(row=row, column=j, value=t)
        c.font = BLACKB; c.fill = HDR; c.border = BORDER; c.alignment = CEN

def main(out="Adjusted_Trial_Balance.xlsx"):
    live = LiveCells()
    ajes = A.build_ajes()
    tc = A.transposition_correction()
    corrected = A.corrected_tb()
    adj = A.adjusted_tb()
    eff = A.aje_effects(ajes)
    accts = list(A.COA.keys())

    wb = openpyxl.Workbook()

    # ================= Unadjusted TB =================
    wu = wb.active; wu.title = "Unadjusted TB"
    wu["A1"] = "Meridian Technical Services LLC  |  Unadjusted Trial Balance  |  June 30, 2025"
    wu["A1"].font = BLACKB
    hdr(wu, 3, ["Account", "Debit", "Credit"])
    r = 4
    unadj_row = {}
    for a in accts:
        v = A.UNADJUSTED.get(a, 0)
        wu.cell(row=r, column=1, value=a).font = BLACK
        if v > 0: wu.cell(row=r, column=2, value=v).font = BLACK; wu.cell(row=r,column=2).number_format="#,##0"
        elif v < 0: wu.cell(row=r, column=3, value=-v).font = BLACK; wu.cell(row=r,column=3).number_format="#,##0"
        for c in range(1,4): wu.cell(row=r,column=c).border=BORDER
        wu.cell(row=r,column=1).alignment=LEF
        unadj_row[a] = r
        r += 1
    tot = r
    live.set(wu, f"B{tot}", f"=SUM(B4:B{r-1})", sum(v for v in A.UNADJUSTED.values() if v>0),
             kind="int", number_format="#,##0", font=BLACKB, border=BORDER)
    live.set(wu, f"C{tot}", f"=SUM(C4:C{r-1})", -sum(v for v in A.UNADJUSTED.values() if v<0),
             kind="int", number_format="#,##0", font=BLACKB, border=BORDER)
    wu.cell(row=tot, column=1, value="Totals (out of balance)").font = BLACKB
    wu.cell(row=tot+2, column=1, value=f"Out of balance by {tc['diff']:,} (AR ledger {tc['from_bal']:,} vs "
            f"aging {tc['to_bal']:,}; divisible by 9 = transposition).").font = BLACK
    for j,w in enumerate([28,14,14],1): wu.column_dimensions[get_column_letter(j)].width=w

    # ================= Adjusting Entries =================
    wae = wb.create_sheet("Adjusting Entries")
    wae["A1"] = "Adjusting journal entries and the trial-balance correction"; wae["A1"].font = BLACKB
    wae["A3"] = ("Step 1 (not a journal entry): restate Accounts Receivable from "
                 f"{tc['from_bal']:,} to the aging total {tc['to_bal']:,}. This is a transcription "
                 "correction, not a posting.")
    wae["A3"].font = BLACK
    hdr(wae, 5, ["Entry", "Account", "Debit", "Credit", "Balanced?"])
    r = 6
    for title, lines in ajes:
        first = r
        for k, (acct, dr, cr) in enumerate(lines):
            wae.cell(row=r, column=1, value=title if k==0 else "").font = BLACK
            wae.cell(row=r, column=2, value=acct).font = BLACK
            if dr: wae.cell(row=r, column=3, value=dr).font=BLACK; wae.cell(row=r,column=3).number_format="#,##0"
            if cr: wae.cell(row=r, column=4, value=cr).font=BLACK; wae.cell(row=r,column=4).number_format="#,##0"
            for c in range(1,6): wae.cell(row=r,column=c).border=BORDER
            wae.cell(row=r,column=1).alignment=LEF; wae.cell(row=r,column=2).alignment=LEF
            r += 1
        # balanced check live formula on the first row of the entry
        live.set(wae, f"E{first}", f'=IF(SUM(C{first}:C{r-1})=SUM(D{first}:D{r-1}),"OK","ERR")',
                 "OK", kind="str", font=BLACK, border=BORDER, align=CEN)
    for j,w in enumerate([40,26,12,12,10],1): wae.column_dimensions[get_column_letter(j)].width=w

    # ================= Adjusted TB =================
    wa = wb.create_sheet("Adjusted TB")
    wa["A1"] = "Adjusted Trial Balance  |  June 30, 2025"; wa["A1"].font = BLACKB
    hdr(wa, 3, ["Account", "Unadjusted (corrected)", "Net adjustment", "Adjusted Debit", "Adjusted Credit"])
    r = 4
    adj_first = r
    for a in accts:
        cb = corrected[a]                 # corrected unadjusted (AR restated)
        e = eff.get(a, 0)
        av = adj.get(a, 0)
        wa.cell(row=r, column=1, value=a).font = BLACK; wa.cell(row=r,column=1).alignment=LEF
        # corrected unadjusted (value)
        wa.cell(row=r, column=2, value=cb).font = BLACK; wa.cell(row=r,column=2).number_format="#,##0;(#,##0)"
        # net adjustment (value from AJEs)
        wa.cell(row=r, column=3, value=e).font = BLACK; wa.cell(row=r,column=3).number_format="#,##0;(#,##0)"
        # adjusted debit/credit as live formulas off B+C
        signed = f"(B{r}+C{r})"
        live.set(wa, f"D{r}", f"=IF({signed}>0,{signed},0)", av if av>0 else 0,
                 kind="int", number_format="#,##0", font=BLACK, border=BORDER)
        live.set(wa, f"E{r}", f"=IF({signed}<0,-{signed},0)", -av if av<0 else 0,
                 kind="int", number_format="#,##0", font=BLACK, border=BORDER)
        for c in range(1,6): wa.cell(row=r,column=c).border=BORDER
        r += 1
    tot = r
    wa.cell(row=tot, column=1, value="Totals").font = BLACKB
    drA = sum(v for v in adj.values() if v>0); crA = -sum(v for v in adj.values() if v<0)
    live.set(wa, f"D{tot}", f"=SUM(D{adj_first}:D{r-1})", drA, kind="int",
             number_format="#,##0", font=BLACKB, border=BORDER)
    live.set(wa, f"E{tot}", f"=SUM(E{adj_first}:E{r-1})", crA, kind="int",
             number_format="#,##0", font=BLACKB, border=BORDER)
    # balance check
    wa.cell(row=tot+2, column=1, value="Debits = Credits?").font = BLACKB
    live.set(wa, f"B{tot+2}", f'=IF(D{tot}=E{tot},"BALANCED","OUT")', "BALANCED",
             kind="str", font=BLACKB, border=BORDER, align=CEN)
    from openpyxl.formatting.rule import CellIsRule
    wa.conditional_formatting.add(f"B{tot+2}",
        CellIsRule(operator="equal", formula=['"BALANCED"'], fill=PatternFill("solid", fgColor="C6EFCE")))
    # net income
    wa.cell(row=tot+3, column=1, value="Adjusted net income").font = BLACKB
    # NI = revenue - expenses, live off adjusted balances
    rev_rows = [adj_first + accts.index(a) for a in accts if A.COA[a]=="R"]
    exp_rows = [adj_first + accts.index(a) for a in accts if A.COA[a]=="X"]
    rev_terms = "+".join(f"E{rr}" for rr in rev_rows)
    exp_terms = "+".join(f"D{rr}" for rr in exp_rows)
    live.set(wa, f"B{tot+3}", f"=({rev_terms})-({exp_terms})", A.net_income(adj),
             kind="int", number_format="#,##0", font=BLACKB, border=BORDER)
    for j,w in enumerate([28,20,16,16,16],1): wa.column_dimensions[get_column_letter(j)].width=w

    # ================= Bank Rec =================
    wb_ = wb.create_sheet("Bank Rec")
    wb_["A1"] = "Bank reconciliation, June 2025"; wb_["A1"].font = BLACKB
    b = A.BANK
    rows_bank = [
        ("Balance per bank", b["balance_per_bank"]),
        ("Add: deposit in transit", b["deposit_in_transit"]),
        ("Less: outstanding checks", -b["outstanding_checks"]),
    ]
    hdr(wb_, 3, ["Bank side", "Amount"])
    r = 4
    for name, amt in rows_bank:
        wb_.cell(row=r,column=1,value=name).font=BLACK; wb_.cell(row=r,column=1).alignment=LEF
        wb_.cell(row=r,column=2,value=amt).font=BLACK; wb_.cell(row=r,column=2).number_format="#,##0;(#,##0)"
        for c in range(1,3): wb_.cell(row=r,column=c).border=BORDER
        r += 1
    live.set(wb_, f"B{r}", f"=SUM(B4:B{r-1})", b["balance_per_bank"]+b["deposit_in_transit"]-b["outstanding_checks"],
             kind="int", number_format="#,##0", font=BLACKB, border=BORDER)
    wb_.cell(row=r,column=1,value="Adjusted bank balance").font=BLACKB; wb_.cell(row=r,column=1).border=BORDER
    # book side
    rr = r + 2
    hdr(wb_, rr, ["Book side", "Amount"]); rr += 1
    book_rows = [
        ("Balance per books (cash)", A.UNADJUSTED["1000 Cash"]),
        ("Less: bank service fee", -b["bank_service_fee"]),
        ("Less: NSF check", -b["nsf_check"]),
    ]
    bstart = rr
    for name, amt in book_rows:
        wb_.cell(row=rr,column=1,value=name).font=BLACK; wb_.cell(row=rr,column=1).alignment=LEF
        wb_.cell(row=rr,column=2,value=amt).font=BLACK; wb_.cell(row=rr,column=2).number_format="#,##0;(#,##0)"
        for c in range(1,3): wb_.cell(row=rr,column=c).border=BORDER
        rr += 1
    live.set(wb_, f"B{rr}", f"=SUM(B{bstart}:B{rr-1})",
             A.UNADJUSTED["1000 Cash"]-b["bank_service_fee"]-b["nsf_check"],
             kind="int", number_format="#,##0", font=BLACKB, border=BORDER)
    wb_.cell(row=rr,column=1,value="Adjusted book cash").font=BLACKB; wb_.cell(row=rr,column=1).border=BORDER
    wb_.cell(row=rr+2,column=1,value="Reconciles?").font=BLACKB
    live.set(wb_, f"B{rr+2}", f'=IF(B{r}=B{rr},"TIES","DIFF")', "TIES",
             kind="str", font=BLACKB, border=BORDER, align=CEN)
    wb_.conditional_formatting.add(f"B{rr+2}",
        CellIsRule(operator="equal", formula=['"TIES"'], fill=PatternFill("solid", fgColor="C6EFCE")))
    wb_.column_dimensions["A"].width = 28; wb_.column_dimensions["B"].width = 14

    wb.move_sheet("Adjusted TB", -(wb.sheetnames.index("Adjusted TB")))
    wb.save(out)
    n, files = live.inject(out)
    normalize_decimals(out, max_dp=6)
    set_excel_fingerprint(out, application="Microsoft Excel", creator="Accounting")
    print(f"wrote {out}: {n} live cells; adjusted TB Dr=Cr={drA:,}; NI={A.net_income(adj):,}")

if __name__ == "__main__":
    main()

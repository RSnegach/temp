# -*- coding: utf-8 -*-
"""
Build the raw General Ledger workbook for the hardened accounting retest:
  Sheet 'General Ledger' : opening balances + dated month activity (foots to the
                           unadjusted TB; the analyst must sum per account).
  Sheet 'Subledgers'     : AR aging total, inventory stock count, notes with traps.
  Sheet 'Bank Statement' : bank balance and items.
No adjusted numbers here; this is the raw input. Black text, neutral header.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import accounting as A
import accounting_hard as H
from xlsx_live import normalize_decimals, set_excel_fingerprint

thin = Side(style="thin", color="000000")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HDR = PatternFill("solid", fgColor="D9D9D9")
BLACK = Font(color="000000"); BLACKB = Font(color="000000", bold=True)
CEN = Alignment(horizontal="center"); LEF = Alignment(horizontal="left")

def hdr(ws, row, labels):
    for j, t in enumerate(labels, 1):
        c = ws.cell(row=row, column=j, value=t)
        c.font = BLACKB; c.fill = HDR; c.border = BORDER; c.alignment = CEN

def main(out="General_Ledger_June.xlsx"):
    wb = openpyxl.Workbook()

    # ---------- General Ledger ----------
    ws = wb.active; ws.title = "General Ledger"
    ws["A1"] = "Meridian Technical Services LLC  |  General Ledger detail  |  Month ended June 30, 2025"
    ws["A1"].font = BLACKB
    hdr(ws, 3, ["Date", "Ref", "Account", "Memo", "Debit", "Credit"])
    r = 4
    # opening balances as one compound entry (signed -> dr/cr columns)
    op = H.opening_balances()
    ws.cell(row=r, column=1, value="06-01").font = BLACK
    ws.cell(row=r, column=2, value="OB").font = BLACK
    ws.cell(row=r, column=3, value="Opening balances (all accounts)").font = BLACK
    ws.cell(row=r, column=4, value="carried forward from May close").font = BLACK
    for c in range(1, 7): ws.cell(row=r, column=c).border = BORDER
    r += 1
    for acct in A.COA:
        v = op.get(acct, 0)
        if v == 0:
            continue
        dr = v if v > 0 else 0
        cr = -v if v < 0 else 0
        ws.cell(row=r, column=3, value=acct).font = BLACK
        ws.cell(row=r, column=4, value="opening").font = BLACK
        if dr: ws.cell(row=r, column=5, value=dr).font = BLACK
        if cr: ws.cell(row=r, column=6, value=cr).font = BLACK
        for c in range(1, 7): ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=5).number_format = "#,##0"
        ws.cell(row=r, column=6).number_format = "#,##0"
        r += 1
    # month activity
    ref = 100
    for date, memo, lines in H.MONTH_ACTIVITY:
        for acct, dr, cr in lines:
            ws.cell(row=r, column=1, value=date).font = BLACK
            ws.cell(row=r, column=2, value=f"J{ref}").font = BLACK
            ws.cell(row=r, column=3, value=acct).font = BLACK
            ws.cell(row=r, column=4, value=memo).font = BLACK
            if dr: ws.cell(row=r, column=5, value=dr).font = BLACK
            if cr: ws.cell(row=r, column=6, value=cr).font = BLACK
            for c in range(1, 7): ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=5).number_format = "#,##0"
            ws.cell(row=r, column=6).number_format = "#,##0"
            r += 1
        ref += 1
    for j, w in enumerate([9, 7, 26, 34, 12, 12], 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    # ---------- Subledgers ----------
    ws2 = wb.create_sheet("Subledgers and Counts")
    ws2["A1"] = "Supporting records at June 30, 2025"; ws2["A1"].font = BLACKB
    hdr(ws2, 3, ["Record", "Amount", "Note"])
    subs = [
        ("Accounts Receivable aging total", H.AR_SUBLEDGER,
         "Sum of the customer aging schedule."),
        ("Inventory physical count", H.INVENTORY_COUNT,
         "Warehouse stock count at period end; variance to GL is under review by operations."),
        ("Prepaid insurance policy", 18000, "12-month policy, 4 months elapsed at period end."),
        ("Equipment depreciation", 24000, "Annual straight-line; book one month for June."),
        ("Wages earned, unpaid at 6/30", 7600, "Payroll period cut off mid-week."),
        ("Unearned revenue earned in June", 12000, "Portion of the customer advance now delivered."),
        ("Supplies misposted to Rent", 6000, "A supplies purchase was recorded to Rent Expense in error."),
    ]
    rr = 4
    for name, amt, note in subs:
        ws2.cell(row=rr, column=1, value=name).font = BLACK
        c = ws2.cell(row=rr, column=2, value=amt); c.font = BLACK; c.number_format = "#,##0"
        ws2.cell(row=rr, column=3, value=note).font = BLACK
        for c in range(1, 4): ws2.cell(row=rr, column=c).border = BORDER
        ws2.cell(row=rr, column=1).alignment = LEF; ws2.cell(row=rr, column=3).alignment = LEF
        rr += 1
    ws2.cell(row=rr+1, column=1, value="Reminder: June rent of 6,000 is paid monthly and is recorded in the ledger this month.").font = BLACK
    for j, w in enumerate([34, 12, 62], 1):
        ws2.column_dimensions[get_column_letter(j)].width = w

    # ---------- Bank Statement ----------
    ws3 = wb.create_sheet("Bank Statement")
    ws3["A1"] = "Bank statement summary, June 2025"; ws3["A1"].font = BLACKB
    hdr(ws3, 3, ["Item", "Amount", "Note"])
    items = [
        ("Balance per bank", A.BANK["balance_per_bank"], "Closing balance on the June statement."),
        ("Deposit in transit", A.BANK["deposit_in_transit"], "Recorded on books 6/30; not yet on statement."),
        ("Outstanding checks", A.BANK["outstanding_checks"], "Issued, not yet cleared."),
        ("Bank service fee", A.BANK["bank_service_fee"], "Charged by bank; not yet on the books."),
        ("NSF check returned", A.BANK["nsf_check"], "Customer check deposited then returned; not yet on the books."),
    ]
    rr = 4
    for name, amt, note in items:
        ws3.cell(row=rr, column=1, value=name).font = BLACK
        c = ws3.cell(row=rr, column=2, value=amt); c.font = BLACK; c.number_format = "#,##0"
        ws3.cell(row=rr, column=3, value=note).font = BLACK
        for c in range(1, 4): ws3.cell(row=rr, column=c).border = BORDER
        ws3.cell(row=rr, column=1).alignment = LEF; ws3.cell(row=rr, column=3).alignment = LEF
        rr += 1
    for j, w in enumerate([22, 12, 56], 1):
        ws3.column_dimensions[get_column_letter(j)].width = w

    wb.save(out)
    normalize_decimals(out, max_dp=6)
    set_excel_fingerprint(out, application="Microsoft Excel", creator="Accounting")
    print("wrote", out)

if __name__ == "__main__":
    main()

# -*- coding: utf-8 -*-
"""
Add machine-readable schedules to the BESS golden so the connectivity and counts
are directly verifiable (not only encoded as colored diagram cells). Adds, for
each build config, a Block_Inventory sheet and a Connection_List sheet generated
from topology.py. The two visual diagram sheets are kept as-is.
Headers: black by default, one accent sheet (never navy #15324B).
"""
import sys
sys.path.insert(0,"scripts")
import topology as T
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

GOLD="golden/bess_block_diagram.xlsx"
BLACK="000000"; WHITE="FFFFFF"; ACCENT="6B7B8C"   # slate for one sheet
thin=Side(style="thin",color="BBBBBB"); B=Border(left=thin,right=thin,top=thin,bottom=thin)

def hdr(ws,row,headers,fill=BLACK):
    for i,h in enumerate(headers):
        c=ws.cell(row,1+i,h)
        c.fill=PatternFill("solid",fgColor=fill)
        c.font=Font(name="Calibri",size=10,bold=True,color=WHITE)
        c.alignment=Alignment(horizontal="center",vertical="center")
        c.border=B

TYPE_NAME=T.TYPES  # key -> label

wb=load_workbook(GOLD)
# remove any prior register sheets so this is idempotent
for s in ["Std_Inventory","Std_Connections","Ext_Inventory","Ext_Connections"]:
    if s in wb.sheetnames: del wb[s]

def add_config(nblocks, inv_title, con_title, accent):
    blocks, edges = T.build(nblocks)
    # inventory
    inv=wb.create_sheet(inv_title)
    inv["A1"]=f"Block Inventory ({nblocks} power blocks)"
    inv["A1"].font=Font(name="Calibri",size=11,bold=True,color="333333")
    hdr(inv,2,["Block type","Tag prefix","Count"], fill=accent)
    from collections import Counter, OrderedDict
    order=["SC","AUXX","MVX","MTR","BUS","CMB","PCS","DCC","RK","BMS"]
    cnt=Counter(b["type"] for b in blocks)
    r=3
    for t in order:
        if cnt.get(t):
            inv.cell(r,1,TYPE_NAME[t]); inv.cell(r,2,t); inv.cell(r,3,cnt[t])
            for c in range(1,4): inv.cell(r,c).border=B; inv.cell(r,c).font=Font(name="Calibri",size=10)
            r+=1
    inv.cell(r,1,"TOTAL BLOCKS").font=Font(name="Calibri",size=10,bold=True)
    inv.cell(r,3,len(blocks)).font=Font(name="Calibri",size=10,bold=True)
    for c in range(1,4): inv.cell(r,c).border=B
    for col,w in zip("ABC",[26,10,8]): inv.column_dimensions[col].width=w
    inv.freeze_panes="A3"

    # connection list
    con=wb.create_sheet(con_title)
    con["A1"]=f"Connection List ({nblocks} power blocks)"
    con["A1"].font=Font(name="Calibri",size=11,bold=True,color="333333")
    hdr(con,2,["#","From","To","Harness"])
    # sort edges by harness then from for readability
    horder={"DC-PWR":0,"AC-PWR":1,"COMMS":2,"AUX-24V":3}
    es=sorted(edges, key=lambda e:(horder.get(e["harness"],9), e["a"], e["b"]))
    for i,e in enumerate(es,1):
        rr=2+i
        con.cell(rr,1,i); con.cell(rr,2,e["a"]); con.cell(rr,3,e["b"]); con.cell(rr,4,e["harness"])
        for c in range(1,5): con.cell(rr,c).border=B; con.cell(rr,c).font=Font(name="Calibri",size=9)
    # per-harness count summary below
    from collections import Counter as C2
    hc=C2(e["harness"] for e in edges)
    base=2+len(es)+2
    con.cell(base,2,"Edge counts by harness").font=Font(name="Calibri",size=10,bold=True)
    for j,h in enumerate(["DC-PWR","AC-PWR","COMMS","AUX-24V"]):
        con.cell(base+1+j,2,h); con.cell(base+1+j,3,hc.get(h,0))
        for c in (2,3): con.cell(base+1+j,c).border=B; con.cell(base+1+j,c).font=Font(name="Calibri",size=9)
    con.cell(base+5,2,"TOTAL").font=Font(name="Calibri",size=10,bold=True)
    con.cell(base+5,3,len(edges)).font=Font(name="Calibri",size=10,bold=True)
    for c in (2,3): con.cell(base+5,c).border=B
    for col,w in zip("ABCD",[5,12,12,10]): con.column_dimensions[col].width=w
    con.freeze_panes="A3"
    return len(blocks), len(edges), dict(hc)

s1=add_config(T.STANDARD_BLOCKS,"Std_Inventory","Std_Connections",BLACK)
s2=add_config(T.EXTENDED_BLOCKS,"Ext_Inventory","Ext_Connections",ACCENT)

# order sheets: diagrams first, then schedules
wb.properties.creator="Electrical Drafting"; wb.properties.lastModifiedBy="Electrical Drafting"
wb.save(GOLD)
print("standard:", s1)
print("extended:", s2)
print("sheets:", wb.sheetnames)

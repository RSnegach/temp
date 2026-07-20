# -*- coding: utf-8 -*-
"""
Build the golden obligation-register deliverable: Obligation_Register.xlsx.

The register is a live-formula workbook. Due dates are computed with real Excel
date functions off an anchor-dates block and a holiday range:
  - calendar-day deadlines: anchor + N, then rolled with WORKDAY.INTL(...-1,1,hol)
    if they land on a weekend/holiday (roll forward to next business day).
  - business-day deadlines: WORKDAY.INTL(anchor, N, weekend, holidays).
  - "N business days before" an anchor: WORKDAY.INTL(anchor, -N, weekend, holidays).
  - monthly 5th: DATE(year,m,5) rolled forward.
  - chained: uses the SOW sign-off due-date cell as the anchor.
Cached values come from contract.py so the golden ties out. Black text, neutral
header. Weekend code "0000011" = Sat/Sun weekend for WORKDAY.INTL.
"""
import datetime as dt
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import contract as K
from xlsx_live import LiveCells, normalize_decimals, set_excel_fingerprint

thin = Side(style="thin", color="000000")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HDR = PatternFill("solid", fgColor="D9D9D9")
BLACK = Font(color="000000"); BLACKB = Font(color="000000", bold=True)
CEN = Alignment(horizontal="center"); LEF = Alignment(horizontal="left")
# Use plain WORKDAY (native since Excel 2007, Sat/Sun weekend, no _xlfn prefix
# needed). WORKDAY.INTL is a post-2007 function that openpyxl writes without the
# _xlfn prefix -> #NAME? in Excel, so avoid it.

def hdr(ws, row, labels):
    for j, t in enumerate(labels, 1):
        c = ws.cell(row=row, column=j, value=t)
        c.font = BLACKB; c.fill = HDR; c.border = BORDER; c.alignment = CEN

def main(out="Obligation_Register.xlsx"):
    live = LiveCells()
    rows = K.register()
    wb = openpyxl.Workbook()

    # ---------------- Anchors ----------------
    wan = wb.active; wan.title = "Anchors and Holidays"
    wan["A1"] = "Key dates, facts, and the holiday calendar"; wan["A1"].font = BLACKB
    wan["A3"] = "Effective Date"; wan["A3"].font = BLACK
    wan["B3"] = K.EFFECTIVE_DATE; wan["B3"].number_format = "yyyy-mm-dd"; wan["B3"].font = BLACK
    wan["A4"] = "Amendment No.1 Date"; wan["A4"].font = BLACK
    wan["B4"] = K.AMENDMENT_DATE; wan["B4"].number_format = "yyyy-mm-dd"; wan["B4"].font = BLACK
    wan["A5"] = "Total annual fees"; wan["A5"].font = BLACK
    wan["B5"] = K.FACTS["total_annual_fees"]; wan["B5"].number_format = "#,##0"; wan["B5"].font = BLACK
    wan["A6"] = "Processes personal data"; wan["A6"].font = BLACK
    wan["B6"] = "No" if not K.FACTS["processes_personal_data"] else "Yes"; wan["B6"].font = BLACK
    wan["A7"] = "Data classification (Exhibit B)"; wan["A7"].font = BLACK
    wan["B7"] = K.FACTS["data_classification"]; wan["B7"].font = BLACK
    wan["D2"] = "Holiday calendar"; wan["D2"].font = BLACKB
    hols = sorted(K.HOLIDAYS)
    for i, h in enumerate(hols):
        wan.cell(row=3+i, column=4, value=h).number_format = "yyyy-mm-dd"
        wan.cell(row=3+i, column=4).font = BLACK
        wan.cell(row=3+i, column=5, value=K.HOLIDAYS[h]).font = BLACK
    hol_range = f"$D$3:$D${2+len(hols)}"
    wan.column_dimensions["A"].width = 26; wan.column_dimensions["B"].width = 14
    wan.column_dimensions["D"].width = 13; wan.column_dimensions["E"].width = 22

    # ---------------- Register ----------------
    ws = wb.create_sheet("Obligation Register")
    ws["A1"] = "Contract MER-2025-0087  |  Obligation deadline calendar  |  CY2025"
    ws["A1"].font = BLACKB
    hdr(ws, 3, ["#", "ID", "Obligation", "Party", "Basis", "Due date", "Weekday"])

    A = "'Anchors and Holidays'!"
    EFF = f"{A}$B$3"; AMD = f"{A}$B$4"; HOL = f"{A}{hol_range}"

    # map each row id to a formula that reproduces its due date
    def formula_for(rid, base_id):
        """base_id strips the .n occurrence suffix."""
        if base_id == "O1":   # 10 cal days, roll fwd
            return f"=WORKDAY({EFF}+10-1,1,{HOL})"
        if base_id == "O2":   # 5 business days after
            return f"=WORKDAY({EFF},5,{HOL})"
        if base_id == "O3":   # 30 cal days, roll
            return f"=WORKDAY({EFF}+30-1,1,{HOL})"
        if base_id == "O6":   # 15 cal days, roll
            return f"=WORKDAY({EFF}+15-1,1,{HOL})"
        if base_id == "O5":   # fixed Jul 31, roll
            return f"=WORKDAY(DATE(2025,7,31)-1,1,{HOL})"
        if base_id == "O7A":  # 90 business days before Dec 31
            return f"=WORKDAY(DATE(2025,12,31),-90,{HOL})"
        if base_id == "O9":   # 45 cal days after amendment, roll
            return f"=WORKDAY({AMD}+45-1,1,{HOL})"
        if base_id == "O11":  # chained: 20 business days after O3 due
            return None       # filled after we know O3's row
        if base_id == "O12":  # 75 cal days after eff, roll
            return f"=WORKDAY({EFF}+75-1,1,{HOL})"
        if base_id == "O4":   # 3 business days before quarter end
            return "QBR"      # handled per-occurrence
        if base_id == "O8":   # monthly 5th, roll
            return "MON"
        return None

    # quarter ends and month per occurrence, from the row title/order
    q_ends = {1: "DATE(2025,3,31)", 2: "DATE(2025,6,30)", 3: "DATE(2025,9,30)", 4: "DATE(2025,12,31)"}
    o3_row = None
    r = 4
    # first pass to find O3 row for chaining
    for i, row in enumerate(rows):
        if row["id"] == "O3":
            o3_row = 4 + i
    for i, row in enumerate(rows):
        rr = 4 + i
        base = row["id"].split(".")[0]
        ws.cell(row=rr, column=1, value=i+1).font = BLACK
        ws.cell(row=rr, column=2, value=row["id"]).font = BLACK
        ws.cell(row=rr, column=3, value=row["title"]).font = BLACK
        ws.cell(row=rr, column=4, value=row["party"]).font = BLACK
        ws.cell(row=rr, column=5, value=row["basis"]).font = BLACK
        f = formula_for(row["id"], base)
        if f == "QBR":
            occ = int(row["id"].split(".")[1])
            f = f"=WORKDAY({q_ends[occ]},-3,{HOL})"
        elif f == "MON":
            occ = int(row["id"].split(".")[1])
            month = occ + 1   # O8.1 = Feb
            f = f"=WORKDAY(DATE(2025,{month},5)-1,1,{HOL})"
        elif base == "O11":
            f = f"=WORKDAY(F{o3_row},20,{HOL})"
        live.set(ws, f"F{rr}", f, row["due"], kind="date", number_format="yyyy-mm-dd",
                 font=BLACK, border=BORDER, align=CEN)
        # weekday formula
        live.set(ws, f"G{rr}", f'=TEXT(F{rr},"ddd")', row["due"].strftime("%a"),
                 kind="str", font=BLACK, border=BORDER, align=CEN)
        for c in range(1, 8): ws.cell(row=rr, column=c).border = BORDER
        ws.cell(row=rr, column=3).alignment = LEF; ws.cell(row=rr, column=5).alignment = LEF
    # count
    cnt_r = 4 + len(rows) + 1
    ws.cell(row=cnt_r, column=5, value="Total active dated rows").font = BLACKB
    live.set(ws, f"F{cnt_r}", f"=COUNTA(B4:B{3+len(rows)})", len(rows), kind="int",
             number_format="0", font=BLACKB, border=BORDER, align=CEN)

    for j, w in enumerate([5, 7, 34, 9, 60, 13, 9], 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    # ---------------- Dropped ----------------
    wd = wb.create_sheet("Dropped")
    wd["A1"] = "Obligations not on the calendar (with reason)"; wd["A1"].font = BLACKB
    hdr(wd, 3, ["ID", "Obligation", "Reason"])
    dropped = [
        ("O7", "Non-renewal notice (original 60 business days)",
         "Superseded: Amendment No.1 deletes Section 7.2 in its entirety and replaces the 60 with 90 business days (see O7A)."),
        ("O10", "Data-protection breach drill",
         "Condition not met: applies only if the Provider Processes Personal Data; the recital states it does not."),
    ]
    for i, (oid, title, reason) in enumerate(dropped):
        rr = 4 + i
        wd.cell(row=rr, column=1, value=oid).font = BLACK
        wd.cell(row=rr, column=2, value=title).font = BLACK
        wd.cell(row=rr, column=3, value=reason).font = BLACK
        for c in range(1, 4): wd.cell(row=rr, column=c).border = BORDER
        wd.cell(row=rr, column=2).alignment = LEF; wd.cell(row=rr, column=3).alignment = LEF
    for j, w in enumerate([7, 40, 78], 1):
        wd.column_dimensions[get_column_letter(j)].width = w

    # order sheets: Register first
    wb.move_sheet("Obligation Register", -1)
    wb.save(out)
    n, files = live.inject(out)
    normalize_decimals(out, max_dp=6)
    set_excel_fingerprint(out, application="Microsoft Excel", creator="Legal Operations")
    print(f"wrote {out}: {n} live cells; {len(rows)} obligations, {len(dropped)} dropped")

if __name__ == "__main__":
    main()

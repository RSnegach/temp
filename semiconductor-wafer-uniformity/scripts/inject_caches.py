# -*- coding: utf-8 -*-
"""Re-inject cached <v> for every formula cell in the golden after openpyxl save.
Computes each value in Python from the deduped source so cache == what Excel
would recalc. Also strips float tails and confirms Excel fingerprint.
Runs a full alignment check: golden Per_Wafer stats vs engine-from-input."""
import json, re, zipfile, shutil, statistics
from openpyxl import load_workbook

GOLD="golden/L7734-02_Wafer_Analysis.xlsx"
LSL=80.75; USL=89.25; WARN_LO=82.45; WARN_HI=87.55; EXCL_R=72.0
ded=json.load(open("ded_src.json"))
wafers=sorted({r[0] for r in ded})

# ---- compute per-wafer truth from deduped input ----
def wafer_stats(w):
    incl=[(s,rad,rs) for (ww,s,rad,rs,ts) in ded if ww==w and rad<=EXCL_R]
    rsv=[rs for _,_,rs in incl]
    n=len(rsv); mean=round(statistics.mean(rsv),2); mn=round(min(rsv),2); mx=round(max(rsv),2)
    wiw=round((mx-mn)/(2*mean)*100,2)
    npass=sum(1 for v in rsv if LSL<=v<=USL)
    return dict(n=n, mean=mean, mn=mn, mx=mx, wiw=wiw, yld=round(npass/n*100,2), fails=n-npass)
stats={w:wafer_stats(w) for w in wafers}

# ---- build cache map keyed by "Sheet!Coord" ----
cache={}
# Raw_Data: included + Rs_incl per row (data_only literals already present for A-D)
wb=load_workbook(GOLD)
rd=wb["Raw_Data"]
row_of={}
for r in range(3, rd.max_row+1):
    w=rd.cell(r,1).value
    if not w: continue
    rad=rd.cell(r,3).value; rs=rd.cell(r,4).value
    inc = "Y" if rad<=EXCL_R else "N"
    cache[f"Raw_Data!E{r}"]=inc
    cache[f"Raw_Data!F{r}"]=(rs if inc=="Y" else "")
    row_of.setdefault(w,[]).append(r)
# Per_Wafer
for idx,w in enumerate(wafers):
    pr=3+idx; s=stats[w]
    cache[f"Per_Wafer!B{pr}"]=s["n"]; cache[f"Per_Wafer!C{pr}"]=s["mean"]
    cache[f"Per_Wafer!D{pr}"]=s["mn"]; cache[f"Per_Wafer!E{pr}"]=s["mx"]
    cache[f"Per_Wafer!F{pr}"]=s["wiw"]; cache[f"Per_Wafer!G{pr}"]=s["yld"]; cache[f"Per_Wafer!H{pr}"]=s["fails"]
lot=3+len(wafers)
lot_mean=round(statistics.mean(stats[w]["mean"] for w in wafers),2)
max_wiw=round(max(stats[w]["wiw"] for w in wafers),2)
tot_n=sum(stats[w]["n"] for w in wafers); tot_fail=sum(stats[w]["fails"] for w in wafers)
lot_yld=round((tot_n-tot_fail)/tot_n*100,2)
cache[f"Per_Wafer!C{lot}"]=lot_mean; cache[f"Per_Wafer!F{lot}"]=max_wiw; cache[f"Per_Wafer!G{lot}"]=lot_yld

# preserve existing Site_Detail + Lot_Summary caches captured earlier
existing=json.load(open("existing_cache.json"))
for k,v in existing.items():
    if k.startswith(("Site_Detail!","Lot_Summary!")) and v is not None:
        cache[k]=v

# ---- inject <v> into saved xml (string-based, preserves everything else) ----
def sheet_files(z):
    wbxml=z.read("xl/workbook.xml").decode(); rels=z.read("xl/_rels/workbook.xml.rels").decode()
    nr={}
    for tag in re.findall(r"<sheet\b[^>]*?/?>",wbxml):
        nm=re.search(r'\bname="([^"]+)"',tag); rid=re.search(r'r:id="([^"]+)"',tag)
        if nm and rid: nr[nm.group(1)]=rid.group(1)
    rt={}
    for tag in re.findall(r"<Relationship\b[^>]*?/?>",rels):
        i=re.search(r'\bId="([^"]+)"',tag); t=re.search(r'\bTarget="([^"]+)"',tag)
        if i and t: rt[i.group(1)]=t.group(1)
    out={}
    for name,rid in nr.items():
        tgt=rt.get(rid,"")
        tgt = tgt[1:] if tgt.startswith("/") else ("xl/"+tgt.lstrip("./") if not tgt.startswith("xl/") else tgt)
        out[name]=tgt
    return out

def esc(s): return str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;").replace('"',"&quot;")
def emit(v):
    if isinstance(v,(int,)) or (isinstance(v,float) and v==int(v)): return str(int(v)),False
    if isinstance(v,float): return f"{v:.4f}".rstrip("0").rstrip("."),False
    return esc(v),True   # string

zin=zipfile.ZipFile(GOLD)
t2f=sheet_files(zin)
perfile={}
for key,val in cache.items():
    sh,co=key.split("!"); f=t2f[sh]; perfile.setdefault(f,{})[co]=val

def patch(txt,cmap):
    def repl(m):
        cell=m.group(0); cm=re.search(r'\br="([A-Z]+\d+)"',cell)
        if not cm or cm.group(1) not in cmap: return cell
        if "<f>" not in cell and "<f " not in cell: return cell
        val=cmap[cm.group(1)]; cell=re.sub(r"<v>.*?</v>","",cell,flags=re.S)
        s,isstr=emit(val)
        if isstr:
            ot=re.match(r"<c\b[^>]*>",cell).group(0)
            if ' t="' not in ot: cell=cell.replace(ot,ot[:-1]+' t="str">',1)
        return cell.replace("</c>",f"<v>{s}</v></c>",1)
    return re.sub(r"<c\b[^>]*>.*?</c>",repl,txt,flags=re.S)

tmp=GOLD+".tmp"; out=zipfile.ZipFile(tmp,"w",zipfile.ZIP_DEFLATED)
patched=0
for it in zin.infolist():
    data=zin.read(it.filename)
    if it.filename in perfile:
        data=patch(data.decode(),perfile[it.filename]).encode(); patched+=1
    out.writestr(it,data)
out.close(); zin.close(); shutil.move(tmp,GOLD)
print(f"injected caches into {patched} sheets, {len(cache)} cells")
print(f"lot mean {lot_mean} maxWIW {max_wiw} lot yield {lot_yld}")

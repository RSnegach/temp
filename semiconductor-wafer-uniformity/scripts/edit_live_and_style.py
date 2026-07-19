# -*- coding: utf-8 -*-
"""
In-place edits to the committed wafer files (preserving prior manual edits):
  1. Golden Per_Wafer: replace hard-coded stats with live Excel formulas that
     aggregate a new Raw_Data sheet (built from the input's real measurements),
     so the sheet recalculates and stays aligned with the input.
  2. Restyle all header rows to plain black fill with white text, replacing the
     uniform #15324B navy.
"""
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

GOLD="golden/L7734-02_Wafer_Analysis.xlsx"
INP="inputs/L7734-02_Rs_Measurements.xlsx"
LSL=80.75; USL=89.25; EXCL_R=72.0

ded=json.load(open("ded_src.json"))     # [[wafer,site,radius,rs,ts],...] deduped, sorted
wafers=sorted({r[0] for r in ded})

BLACK="000000"; WHITE="FFFFFF"
thin=Side(style="thin",color="BBBBBB")
B=Border(left=thin,right=thin,top=thin,bottom=thin)
# most headers plain black; a couple of sheets get a distinct accent (never #15324B)
ACCENT={"Raw_Data":"6B7B8C", "Wafer_Map":"7A5230"}   # slate grey, muted brown

def restyle_header(ws,row,ncols,start=1,fill=None,font=None):
    fc=fill or BLACK; tc=font or WHITE
    for c in range(start,start+ncols):
        cell=ws.cell(row=row,column=c)
        cell.fill=PatternFill("solid",fgColor=fc)
        f=cell.font
        cell.font=Font(name=f.name or "Calibri", size=f.size or 10, bold=True, color=tc)

# ================= GOLDEN =================
wb=load_workbook(GOLD)

# Raw_Data sheet from input measurements (alignment source for Per_Wafer)
if "Raw_Data" in wb.sheetnames: del wb["Raw_Data"]
rd=wb.create_sheet("Raw_Data")
rd["A1"]="Raw_Data (deduped site measurements from metrology export)"
rd["A1"].font=Font(name="Calibri",size=11,bold=True,color="333333")
for i,h in enumerate(["wafer","site","radius_mm","Rs_ohm_sq","included","Rs_incl"]):
    rd.cell(2,1+i,h)
restyle_header(rd,2,6,fill=ACCENT["Raw_Data"])
r=3; row_of={}
for w,site,rad,rs,ts in ded:
    rd.cell(r,1,w); rd.cell(r,2,site)
    rd.cell(r,3,round(rad,2)).number_format="0.00"
    rd.cell(r,4,round(rs,2)).number_format="0.00"
    rd.cell(r,5,f'=IF(C{r}<={EXCL_R},"Y","N")')
    rd.cell(r,6,f'=IF(E{r}="Y",D{r},"")')
    for c in range(1,7):
        rd.cell(r,c).font=Font(name="Consolas",size=9); rd.cell(r,c).border=B
    row_of.setdefault(w,[]).append(r); r+=1
for col,wd in zip("ABCDEF",[8,7,11,11,10,10]): rd.column_dimensions[col].width=wd
rd.freeze_panes="A3"

# Per_Wafer live formulas (header row2, wafer rows 3.., LOT row after)
pw=wb["Per_Wafer"]
for idx,w in enumerate(wafers):
    prow=3+idx; rows=row_of[w]; r0,r1=min(rows),max(rows)
    incl=f'Raw_Data!$E${r0}:$E${r1}'; rsincl=f'Raw_Data!$F${r0}:$F${r1}'; rsall=f'Raw_Data!$D${r0}:$D${r1}'
    pw.cell(prow,2,f'=COUNTIF({incl},"Y")')
    pw.cell(prow,3,f'=ROUND(AVERAGE({rsincl}),2)')
    pw.cell(prow,4,f'=ROUND(MIN({rsincl}),2)')
    pw.cell(prow,5,f'=ROUND(MAX({rsincl}),2)')
    pw.cell(prow,6,f'=ROUND((E{prow}-D{prow})/(2*C{prow})*100,2)')
    pw.cell(prow,7,f'=ROUND(SUMPRODUCT(({incl}="Y")*({rsall}>={LSL})*({rsall}<={USL}))/B{prow}*100,2)')
    pw.cell(prow,8,f'=B{prow}-SUMPRODUCT(({incl}="Y")*({rsall}>={LSL})*({rsall}<={USL}))')
lot_row=3+len(wafers)
pw.cell(lot_row,3,f'=ROUND(AVERAGE(C3:C{lot_row-1}),2)')
pw.cell(lot_row,6,f'=ROUND(MAX(F3:F{lot_row-1}),2)')
pw.cell(lot_row,7,f'=ROUND((SUM(B3:B{lot_row-1})-SUM(H3:H{lot_row-1}))/SUM(B3:B{lot_row-1})*100,2)')

# plain black headers for the main analytical sheets
for sh,nc in [("Lot_Summary",3),("Per_Wafer",8),("Site_Detail",9)]:
    restyle_header(wb[sh],2,nc)
# Wafer_Map has no column header row; give its title cell the muted-brown accent
wm=wb["Wafer_Map"]
wm["A1"].fill=PatternFill("solid",fgColor=ACCENT["Wafer_Map"])
_f=wm["A1"].font
wm["A1"].font=Font(name=_f.name or "Calibri", size=_f.size or 11, bold=True, color=WHITE)
wb.save(GOLD)
print("golden saved: Raw_Data + live Per_Wafer + varied headers")

# ================= INPUT =================
wbi=load_workbook(INP)
restyle_header(wbi["Measurements"],2,7)
restyle_header(wbi["Site_Map"],2,5)
wbi.save(INP)
print("input saved: black headers")

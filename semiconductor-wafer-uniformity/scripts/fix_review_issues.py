# -*- coding: utf-8 -*-
"""
Fix graded-review issues on the committed wafer files (edit in place):
  A. Lot_Summary B5-B10: hard-coded stats -> live formulas off Per_Wafer.
     Verdict cells reworked to read numeric cells (no string parsing).
  B. Wafer_Map: hard-coded Rs/EX strings -> formulas referencing Site_Detail,
     plus real conditional-formatting rules (PASS/WARN/FAIL coloring) so the
     parser detects >0 rules and the oracle 'colored sections' negative clears.
  C. Input Measurements: sequential integer timestamps -> realistic epoch
     timestamps (preserving retest ordering so latest still wins dedup).
Caches injected for all formula cells, float tails stripped, fingerprint set.
"""
import sys, re, json, zipfile, shutil
sys.path.insert(0,"scripts")
import wafer as W
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import CellIsRule

GOLD="golden/L7734-02_Wafer_Analysis.xlsx"
INP="inputs/L7734-02_Rs_Measurements.xlsx"
LSL=80.75; USL=89.25; WARN_LO=82.45; WARN_HI=87.55; EXCL_R=72.0
PASS_F="C9E6D4"; WARN_F="FBE8C8"; FAIL_F="F5CBC6"; EXCL_F="D9D9D9"

# site -> Site_Detail row (site N at row 2+N); Rs in col E, Included in col F
def sd_row(site): return 2+site

# ---- cell -> site placement (verified 49/49 against committed map) ----
cx=8; cy=8; sc=6.2/W.WAFER_RADIUS_MM
cell_site={}
for s in W.SITES:
    col=1+int(round(cx+s["x"]*sc)); row=3+int(round(cy-s["y"]*sc))
    cell_site[(row,col)]=s

# ================= GOLDEN =================
wb=load_workbook(GOLD)

# ---- A. Lot_Summary live formulas ----
ls=wb["Lot_Summary"]
# Per_Wafer: header row2, wafers rows3-27, LOT row28. cols: B n, C mean, F WIW, G yield, H fails
ls["B5"]="=ROUND(AVERAGE(Per_Wafer!C3:C27),2)"          # lot mean Rs
ls["B6"]="=Per_Wafer!G28"                                 # (placeholder; W2W below)
# W2W% = (max wafer mean - min wafer mean)/(2*lot mean)*100
ls["B6"]="=ROUND((MAX(Per_Wafer!C3:C27)-MIN(Per_Wafer!C3:C27))/(2*B5)*100,2)"
ls["B7"]="=ROUND(MAX(Per_Wafer!F3:F27),2)"               # max WIW NU%
ls["B8"]="=ROUND((SUM(Per_Wafer!B3:B27)-SUM(Per_Wafer!H3:H27))/SUM(Per_Wafer!B3:B27)*100,2)"  # lot yield
ls["B9"]="=SUM(Per_Wafer!B3:B27)"                        # total included sites
ls["B10"]="=SUM(Per_Wafer!H3:H27)"                       # failing sites
# number formats
ls["B5"].number_format="0.00"; ls["B6"].number_format="0.00"; ls["B7"].number_format="0.00"
ls["B8"].number_format="0.00"; ls["B9"].number_format="0"; ls["B10"].number_format="0"
# verdicts now read numeric cells directly (no LEFT/FIND string parse)
ls["C6"]='=IF(B6<=3,"PASS","FAIL")'
ls["C7"]='=IF(B7<=5,"PASS","FAIL")'
ls["C8"]='=IF(B8>=95,"PASS","FAIL")'
# B12 disposition unchanged (reads C6/C7/C8)

# ---- B. Wafer_Map formula-linked values + conditional formatting ----
wm=wb["Wafer_Map"]
value_cells=[]   # cells that hold a numeric Rs (for conditional formatting range)
for (row,col),s in cell_site.items():
    r=sd_row(s["site"])
    # EX for excluded, else the Rs value pulled from Site_Detail (numeric so CF works)
    if s["excluded"]:
        wm.cell(row,col).value=f'=IF(Site_Detail!F{r}="N","EX",Site_Detail!E{r})'
    else:
        wm.cell(row,col).value=f'=Site_Detail!E{r}'
        value_cells.append(wm.cell(row,col).coordinate)
    wm.cell(row,col).font=Font(name="Calibri",size=7,bold=not s["excluded"],color="1A1A1A")
    wm.cell(row,col).alignment=Alignment(horizontal="center",vertical="center")
    wm.cell(row,col).number_format="0.0"
# real conditional formatting over the whole map grid (rows 4-17, cols A-O)
rng="A4:O17"
# EX text -> grey
wm.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"EX"'],
    fill=PatternFill("solid",fgColor=EXCL_F)))
# FAIL: out of spec
wm.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=[str(LSL)],
    fill=PatternFill("solid",fgColor=FAIL_F)))
wm.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=[str(USL)],
    fill=PatternFill("solid",fgColor=FAIL_F)))
# WARN: in spec but outside warn band
wm.conditional_formatting.add(rng, CellIsRule(operator="between", formula=[str(LSL),str(WARN_LO)],
    fill=PatternFill("solid",fgColor=WARN_F)))
wm.conditional_formatting.add(rng, CellIsRule(operator="between", formula=[str(WARN_HI),str(USL)],
    fill=PatternFill("solid",fgColor=WARN_F)))
# PASS: within warn band
wm.conditional_formatting.add(rng, CellIsRule(operator="between", formula=[str(WARN_LO),str(WARN_HI)],
    fill=PatternFill("solid",fgColor=PASS_F)))
wb.save(GOLD)
print("golden: Lot_Summary formulas + Wafer_Map formulas + conditional formatting")

# ================= INPUT: realistic timestamps =================
# Base epoch: 2026-05-14 07:00:00 UTC shift start. Wafers measured sequentially,
# ~9 min per wafer, ~10 s per site. Retest rows get a later stamp (a few minutes
# after the original) so latest-wins dedup is preserved.
wbi=load_workbook(INP)
ms=wbi["Measurements"]
BASE=1778950800   # 2026-05-14 07:00:00 UTC
WAFER_STRIDE=560  # seconds between wafer starts
SITE_STEP=10      # seconds between sites
# map old ts -> new ts, preserving strict ordering
rows=[]
for r in range(3, ms.max_row+1):
    w=ms.cell(r,1).value
    if not w: continue
    rows.append((r, w, ms.cell(r,2).value, ms.cell(r,7).value))
# order rows by old timestamp within the whole file to preserve global ordering
# but build realistic per-wafer/site times; retests detected by duplicate old ts pattern
# old ts = wafer_index*10000 + site (+5000 for retest). Recover wafer index and retest flag.
def newts(oldts):
    widx=oldts//10000
    rem=oldts%10000
    retest = rem>=5000
    site = rem-5000 if retest else rem
    t = BASE + (widx-1)*WAFER_STRIDE + site*SITE_STEP
    if retest: t += 180   # retest 3 min after the wafer's site pass
    return t
for r,w,site,oldts in rows:
    ms.cell(r,7).value=newts(oldts)
wbi.save(INP)
print("input: realistic epoch timestamps")

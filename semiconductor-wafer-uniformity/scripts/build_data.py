# -*- coding: utf-8 -*-
"""Input file 1: Rs measurement dataset (.xlsx). Raw per-site rows incl. retest
duplicates + a site-coordinate reference. No edge-exclusion flag, no stats."""
import wafer as W
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from xlsx_live import LiveCells, normalize_decimals, set_excel_fingerprint

live = LiveCells()

NAVY="15324B"; WHITE="FFFFFF"; LT="EAF0F4"
thin=Side(style="thin",color="C3CDD6"); B=Border(left=thin,right=thin,top=thin,bottom=thin)
HF=Font(name="Calibri",size=10,bold=True,color=WHITE); HFill=PatternFill("solid",fgColor=NAVY)
CF=Font(name="Calibri",size=10,color="1A1A1A"); MONO=Font(name="Consolas",size=9,color="1A1A1A")
TF=Font(name="Calibri",size=14,bold=True,color=NAVY); SF=Font(name="Calibri",size=9,italic=True,color="5A6472")
Z=PatternFill("solid",fgColor=LT)

wb=Workbook()

def hdr(ws,row,headers):
    for i,h in enumerate(headers):
        c=ws.cell(row=row,column=1+i,value=h); c.font=HF; c.fill=HFill; c.border=B
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)

# ---- Sheet 1: Measurements (raw, chronological-ish export with retest rows) ----
ws=wb.active; ws.title="Measurements"
ws["A1"]="Sheet Resistance Metrology Export"; ws["A1"].font=TF
ws["A2"]=("Lot L7734-02  |  Recipe RTA-Anneal-11  |  Tool ARS-4 (4-point probe)  |  Rs in ohm/sq. "
          "One row per site measurement. Rows are as exported from the tool log in measurement order. "
          "Site coordinates are in the Site_Map sheet. timestamp_s is seconds since shift start.")
ws["A2"].font=SF; ws.merge_cells("A2:G2"); ws.row_dimensions[2].height=44
hdr(ws,4,["wafer_id","site","x_mm","y_mm","radius_mm","Rs_ohm_sq","timestamp_s"])
rows=W.measured_rows()
# export order: by wafer, then by timestamp (so retest originals sit before their retest naturally)
rows_sorted=sorted(rows, key=lambda r:(r["wafer"], r["ts"]))
r=5
for row in rows_sorted:
    ws.cell(r,1,f"W{row['wafer']:02d}").font=CF
    ws.cell(r,2,row["site"]).font=MONO
    ws.cell(r,3,row["x"]).font=MONO
    ws.cell(r,4,row["y"]).font=MONO
    ws.cell(r,5,row["r"]).font=MONO
    c=ws.cell(r,6,row["rs"]); c.font=MONO; c.number_format="0.00"
    ws.cell(r,7,row["ts"]).font=MONO
    for col in range(1,8):
        cc=ws.cell(r,col); cc.border=B
        cc.alignment=Alignment(horizontal="center" if col!=1 else "left",vertical="center")
        if (r%2==1): cc.fill=Z
    r+=1
for col,w in zip("ABCDEFG",[9,7,9,9,11,12,12]): ws.column_dimensions[col].width=w
ws.freeze_panes="A5"

# ---- Sheet 2: Site_Map (coordinate reference, 49 sites) ----
ws2=wb.create_sheet("Site_Map")
ws2["A1"]="Site Coordinate Map (49-site)"; ws2["A1"].font=TF
ws2["A2"]=("Nominal probe coordinates for the 49-site pattern, wafer center at origin. Ring 0 is center; "
           "ring index increases outward. Radius is distance from wafer center. See the sampling plan for the "
           "measurement pattern and any exclusion rule.")
ws2["A2"].font=SF; ws2.merge_cells("A2:E2"); ws2.row_dimensions[2].height=30
# radius_mm is a LIVE cell: distance from center computed from x and y.
# Pure geometry, reveals nothing about exclusion or dedup.
ws2["A3"]="radius_mm is computed from the site coordinates (=ROUND(SQRT(x^2+y^2),2))."
ws2["A3"].font=SF; ws2.merge_cells("A3:E3")
hdr(ws2,4,["site","x_mm","y_mm","radius_mm","ring_index"])
r=5
for s in W.SITES:
    ws2.cell(r,1,s["site"]).font=MONO
    ws2.cell(r,2,s["x"]).font=MONO; ws2.cell(r,2).number_format="0.00"
    ws2.cell(r,3,s["y"]).font=MONO; ws2.cell(r,3).number_format="0.00"
    live.set(ws2, f"D{r}", f"=ROUND(SQRT(B{r}^2+C{r}^2),2)", s["r"], kind="num", dp=2,
             number_format="0.00", font=MONO, align=Alignment(horizontal="center"))
    ws2.cell(r,5,s["ring"]).font=MONO
    for col in range(1,6):
        cc=ws2.cell(r,col); cc.border=B; cc.alignment=Alignment(horizontal="center")
        if r%2==1: cc.fill=Z
    r+=1
for col,w in zip("ABCDE",[7,9,9,11,11]): ws2.column_dimensions[col].width=w
ws2.freeze_panes="A5"

wb.properties.creator="Metrology"; wb.properties.title="Rs Measurement Export L7734-02"; wb.properties.lastModifiedBy="Metrology"
out="inputs/L7734-02_Rs_Measurements.xlsx"; wb.save(out)
nc,nf=live.inject(out)
normalize_decimals(out)   # strip binary float tails from all stored values
set_excel_fingerprint(out)
print("saved",out,"rows",len(rows_sorted),f"| live cells {nc} across {nf} sheet(s)")

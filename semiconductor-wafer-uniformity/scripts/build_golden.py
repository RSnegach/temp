# -*- coding: utf-8 -*-
"""Golden deliverable: L7734-02_Wafer_Analysis.xlsx built from wafer.py.
Sheets: Lot_Summary, Per_Wafer, Wafer_Map (grid-art color map for a rep wafer),
Site_Detail (per-site values, one rep wafer). Deterministic; cached values match
the engine. Clean 2-dp display."""
import wafer as W
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from xlsx_live import LiveCells, normalize_decimals, set_excel_fingerprint

live = LiveCells()

NAVY="15324B"; WHITE="FFFFFF"; LT="EAF0F4"
PASS_F="C9E6D4"; WARN_F="FBE8C8"; FAIL_F="F5CBC6"; EXCL_F="D9D9D9"; RESULT="E4EEF3"
thin=Side(style="thin",color="C3CDD6"); B=Border(left=thin,right=thin,top=thin,bottom=thin)
HF=Font(name="Calibri",size=10,bold=True,color=WHITE); HFill=PatternFill("solid",fgColor=NAVY)
CF=Font(name="Calibri",size=10,color="1A1A1A"); MONO=Font(name="Consolas",size=9,color="1A1A1A")
TF=Font(name="Calibri",size=14,bold=True,color=NAVY); SF=Font(name="Calibri",size=9,italic=True,color="5A6472")
BF=Font(name="Calibri",size=10,bold=True,color="1A1A1A"); RF=Font(name="Calibri",size=11,bold=True,color=NAVY)
Z=PatternFill("solid",fgColor=LT)

rows=W.measured_rows(); dd=W.dedup(rows); lot=W.lot_stats(dd)
REP=1   # representative wafer for the map + site detail (a hot wafer with fails)

wb=Workbook()
def hdr(ws,row,headers,start=1):
    for i,h in enumerate(headers):
        c=ws.cell(row=row,column=start+i,value=h); c.font=HF; c.fill=HFill; c.border=B
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)

# ---------- Lot_Summary ----------
ws=wb.active; ws.title="Lot_Summary"
ws["A1"]="Lot L7734-02 - Sheet Resistance Analysis"; ws["A1"].font=TF
ws["A2"]=("RTA anneal Rs uniformity and yield. Statistics computed on included sites only (edge exclusion "
          "r > 72.0 mm applied) after resolving retests (latest timestamp per site). Rs in ohm/sq.")
ws["A2"].font=SF; ws.merge_cells("A2:D2"); ws.row_dimensions[2].height=30
hdr(ws,4,["Metric","Value","Criterion","Verdict"])
crit_lot_yield_pass = lot["lot_yield"]>=95.0
w2w_pass = lot["w2w_pct"]<=3.0
# how many wafers fail the 5% WIW criterion
wiw_fail=[x["wafer"] for x in lot["wafers"] if x["wiw"]>5.0]
rowsL=[
 ("Wafers in lot", f"{W.N_WAFERS}", "", ""),
 ("Sites per wafer (total / included)", f"{W.N_SITES} / {len(W.INCLUDED)}", "16 edge-excluded", ""),
 ("Lot mean Rs (ohm/sq)", f"{lot['lot_mean']:.2f}", f"target {W.RS_TARGET:.0f}", ""),
 ("Wafer-to-wafer W2W (%)", f"{lot['w2w_pct']:.2f}", "<= 3.00 %", "PASS" if w2w_pass else "FAIL"),
 ("Max WIW nonuniformity (%)", f"{max(x['wiw'] for x in lot['wafers']):.2f}",
   "<= 5.00 % each wafer", "PASS" if not wiw_fail else "FAIL"),
 ("Lot yield (included sites in spec)", f"{lot['lot_yield']:.2f} %",
   ">= 95.00 %", "PASS" if crit_lot_yield_pass else "FAIL"),
 ("Total included site measurements", f"{lot['total_sites']}", "25 x 33", ""),
 ("Failing site measurements", f"{lot['total_sites']-lot['total_pass']}", "", ""),
]
r=5
verd_rows={}   # metric label -> row, for live verdict formulas
for metric,val,crit,verd in rowsL:
    ws.cell(r,1,metric).font=BF if r==5 else CF
    ws.cell(r,2,val).font=CF
    ws.cell(r,3,crit).font=SF
    if verd:
        # LIVE verdict: compare the numeric part of this row's value cell to the threshold
        vc=ws.cell(r,4); vc.font=BF
        if metric.startswith("Wafer-to-wafer"):
            live.set(ws,f"D{r}",'=IF(VALUE(LEFT(B{0},FIND(" ",B{0}&" ")-1))<=3,"PASS","FAIL")'.format(r),
                     verd, kind="str", font=BF, align=Alignment(vertical="center"))
        elif metric.startswith("Max WIW"):
            live.set(ws,f"D{r}",f'=IF(VALUE(B{r})<=5,"PASS","FAIL")', verd, kind="str",
                     font=BF, align=Alignment(vertical="center"))
        elif metric.startswith("Lot yield"):
            live.set(ws,f"D{r}",'=IF(VALUE(LEFT(B{0},FIND(" ",B{0}&" ")-1))>=95,"PASS","FAIL")'.format(r),
                     verd, kind="str", font=BF, align=Alignment(vertical="center"))
        if verd=="PASS": vc.fill=PatternFill("solid",fgColor=PASS_F)
        elif verd=="FAIL": vc.fill=PatternFill("solid",fgColor=FAIL_F)
        verd_rows[metric]=r
    for col in range(1,5): ws.cell(r,col).border=B; ws.cell(r,col).alignment=Alignment(vertical="center")
    r+=1
# overall disposition (live): CONTINUE only if all three verdict cells say PASS
overall = "CONTINUE" if (crit_lot_yield_pass and w2w_pass and not wiw_fail) else "HOLD"
dr=r+1
ws.cell(dr,1,"Lot disposition").font=RF
w2w_r=verd_rows["Wafer-to-wafer W2W (%)"]; wiw_r=verd_rows["Max WIW nonuniformity (%)"]; yld_r=verd_rows["Lot yield (included sites in spec)"]
live.set(ws, f"B{dr}",
         f'=IF(AND(D{w2w_r}="PASS",D{wiw_r}="PASS",D{yld_r}="PASS"),"CONTINUE","HOLD")',
         overall, kind="str", font=RF, align=Alignment(vertical="center"),
         fill=PatternFill("solid",fgColor=PASS_F if overall=="CONTINUE" else FAIL_F))
ws.cell(dr,3, "meets all 3 criteria" if overall=="CONTINUE" else "misses >=1 criterion").font=SF
for col in range(1,4): ws.cell(dr,col).border=B
for col,w in zip("ABCD",[34,16,20,10]): ws.column_dimensions[col].width=w
ws.freeze_panes="A5"

# ---------- Per_Wafer ----------
ws2=wb.create_sheet("Per_Wafer")
ws2["A1"]="Per-Wafer Statistics (included sites)"; ws2["A1"].font=TF
ws2["A2"]="WIW NU = (max - min) / (2 x mean) x 100. Yield = included sites within spec / included sites."
ws2["A2"].font=SF; ws2.merge_cells("A2:H2")
hdr(ws2,4,["Wafer","n incl","Mean","Min","Max","WIW NU %","Yield %","Fails"])
r=5
for x in lot["wafers"]:
    ws2.cell(r,1,f"W{x['wafer']:02d}").font=BF
    for col,val,fmt in [(2,x["n"],"0"),(3,x["mean"],"0.00"),(4,x["min"],"0.00"),
                        (5,x["max"],"0.00"),(6,x["wiw"],"0.00"),(7,x["yield_pct"],"0.0"),(8,x["n_fail"],"0")]:
        cc=ws2.cell(r,col,round(val,2) if fmt!="0" else int(val)); cc.font=MONO; cc.number_format=fmt
        cc.alignment=Alignment(horizontal="center")
    # highlight wafers with fails
    if x["n_fail"]>0:
        for col in range(1,9): ws2.cell(r,col).fill=PatternFill("solid",fgColor=WARN_F)
    for col in range(1,9): ws2.cell(r,col).border=B
    r+=1
# lot mean row
ws2.cell(r,1,"LOT").font=BF; ws2.cell(r,1).fill=Z
ws2.cell(r,3,round(lot["lot_mean"],2)).font=BF; ws2.cell(r,3).number_format="0.00"; ws2.cell(r,3).fill=Z
ws2.cell(r,6,round(max(x["wiw"] for x in lot["wafers"]),2)).font=BF; ws2.cell(r,6).number_format="0.00"; ws2.cell(r,6).fill=Z
ws2.cell(r,7,round(lot["lot_yield"],1)).font=BF; ws2.cell(r,7).number_format="0.0"; ws2.cell(r,7).fill=Z
for col in [1,2,3,4,5,6,7,8]: ws2.cell(r,col).border=B;
for c in [2,4,5,8]: ws2.cell(r,c).fill=Z
for col,w in zip("ABCDEFGH",[8,8,9,9,9,11,9,7]): ws2.column_dimensions[col].width=w
ws2.freeze_panes="A5"

# ---------- Wafer_Map (grid-art color map for REP wafer) ----------
ws3=wb.create_sheet("Wafer_Map")
ws3["A1"]=f"Wafer Map - W{REP:02d} (Rs bin by site)"; ws3["A1"].font=TF
ws3["A2"]=("Sites placed by polar coordinate. Green PASS (within 3% of target), amber WARN (in spec, 3-5% off), "
           "red FAIL (out of spec), grey EX = edge-excluded (not in statistics).")
ws3["A2"].font=SF; ws3.merge_cells("A2:M2"); ws3.row_dimensions[2].height=30
# map grid: 15 cols x 15 rows, center at (8,8). scale mm->cell.
GC=15; GR=15; cx=8; cy=8
sc=6.2/W.WAFER_RADIUS_MM  # mm to cells (so 75mm -> ~6.2 cells radius)
off=W.wafer_offset(REP)
# uniform small cells
for col in range(1,GC*2+2): ws3.column_dimensions[get_column_letter(col)].width=5.2
for r in range(4,4+GR+1): ws3.row_dimensions[r].height=22
site_rs={}
for s in W.SITES:
    site_rs[s["site"]]=W.rs_value(s,REP,off)
# place each site
for s in W.SITES:
    gcx = cx + s["x"]*sc
    gcy = cy - s["y"]*sc   # invert y for screen
    col = 1 + int(round(gcx))
    row = 4 + int(round(gcy))
    rs=site_rs[s["site"]]
    if s["excluded"]:
        fill=EXCL_F; txt="EX"
    else:
        b=W.bin_of(rs); fill={"PASS":PASS_F,"WARN":WARN_F,"FAIL":FAIL_F}[b]; txt=f"{rs:.1f}"
    cc=ws3.cell(row,col,txt); cc.fill=PatternFill("solid",fgColor=fill)
    cc.font=Font(name="Calibri",size=7,bold=not s["excluded"],color="1A1A1A")
    cc.alignment=Alignment(horizontal="center",vertical="center")
    cc.border=B
    # site number as a comment-ish adjacent? keep compact: put number tiny in same cell not possible; skip
# legend
lr=4+GR+2
for i,(lab,fill) in enumerate([("PASS",PASS_F),("WARN",WARN_F),("FAIL",FAIL_F),("EX (edge-excluded)",EXCL_F)]):
    cc=ws3.cell(lr+i,2,lab); cc.fill=PatternFill("solid",fgColor=fill); cc.font=CF; cc.border=B
    cc.alignment=Alignment(horizontal="left")
ws3.cell(lr-1,2,f"W{REP:02d} mean {lot['wafers'][REP-1]['mean']:.2f}  WIW {lot['wafers'][REP-1]['wiw']:.2f}%  "
                f"yield {lot['wafers'][REP-1]['yield_pct']:.1f}%").font=BF

# ---------- Site_Detail (rep wafer, per included/excluded site) ----------
# LIVE cells: radius from x/y; Included from radius vs exclusion threshold;
# Bin and In-spec derived from the Rs cell against the spec limits.
ws4=wb.create_sheet("Site_Detail")
ws4["A1"]=f"Site Detail - W{REP:02d}"; ws4["A1"].font=TF
ws4["A2"]=("Per-site Rs for the mapped wafer. Radius, included flag, bin, and in-spec are computed live from the "
           "coordinates and the Rs reading against the spec limits. Excluded sites are not in the wafer statistics.")
ws4["A2"].font=SF; ws4.merge_cells("A2:H2"); ws4.row_dimensions[2].height=28
# spec constants block (so formulas reference cells, not magic numbers).
# Placed at K/L, clear of the helper column I.
ws4["K4"]="constants"; ws4["K4"].font=SF
consts=[("LSL",W.RS_LSL),("USL",W.RS_USL),("warn_lo",W.RS_WARN_LO),
        ("warn_hi",W.RS_WARN_HI),("excl_r",W.EXCL_RADIUS)]
crow={}
for i,(nm,val) in enumerate(consts):
    rr=5+i
    ws4.cell(rr,11,nm).font=MONO
    cc=ws4.cell(rr,12,val); cc.font=MONO; cc.number_format="0.00"
    crow[nm]=f"$L${rr}"
hdr(ws4,4,["Site","x_mm","y_mm","Radius mm","Rs ohm/sq","Included","Bin","In spec","Rs (incl only)"])
r=5
for s in W.SITES:
    rs=site_rs[s["site"]]
    incl = not s["excluded"]
    b = "-" if not incl else W.bin_of(rs)
    inspec = "-" if not incl else ("Y" if (W.RS_LSL<=rs<=W.RS_USL) else "N")
    ws4.cell(r,1,s["site"]).font=MONO
    ws4.cell(r,2,s["x"]).font=MONO; ws4.cell(r,2).number_format="0.00"
    ws4.cell(r,3,s["y"]).font=MONO; ws4.cell(r,3).number_format="0.00"
    # radius (live)
    live.set(ws4, f"D{r}", f"=ROUND(SQRT(B{r}^2+C{r}^2),2)", s["r"], kind="num", dp=2,
             number_format="0.00", font=MONO, align=Alignment(horizontal="center"))
    # Rs literal (raw measurement)
    cc=ws4.cell(r,5,rs); cc.font=MONO; cc.number_format="0.00"; cc.alignment=Alignment(horizontal="center")
    # Included (live): radius <= exclusion threshold
    live.set(ws4, f"F{r}", f'=IF(D{r}<={crow["excl_r"]},"Y","N")', "Y" if incl else "N",
             kind="str", font=MONO, align=Alignment(horizontal="center"))
    # Bin (live): only for included sites
    bin_formula=(f'=IF(F{r}="N","-",'
                 f'IF(OR(E{r}<{crow["LSL"]},E{r}>{crow["USL"]}),"FAIL",'
                 f'IF(OR(E{r}<{crow["warn_lo"]},E{r}>{crow["warn_hi"]}),"WARN","PASS")))')
    bc=live.set(ws4, f"G{r}", bin_formula, b, kind="str", font=MONO,
                align=Alignment(horizontal="center"))
    if incl: bc.fill=PatternFill("solid",fgColor={"PASS":PASS_F,"WARN":WARN_F,"FAIL":FAIL_F}[b])
    else: bc.fill=PatternFill("solid",fgColor=EXCL_F)
    # In spec (live)
    live.set(ws4, f"H{r}", f'=IF(F{r}="N","-",IF(AND(E{r}>={crow["LSL"]},E{r}<={crow["USL"]}),"Y","N"))',
             inspec, kind="str", font=MONO, align=Alignment(horizontal="center"))
    # helper: Rs for included sites only, blank otherwise (feeds MIN/MAX/AVERAGE,
    # which are legacy functions that work in every Excel; avoids MINIFS/MAXIFS
    # which openpyxl writes without the _xlfn prefix Excel needs -> #NAME?)
    live.set(ws4, f"I{r}", f'=IF(F{r}="Y",E{r},"")', (rs if incl else ""),
             kind=("num" if incl else "str"), dp=2, number_format="0.00",
             font=MONO, align=Alignment(horizontal="center"))
    for col in range(1,10): ws4.cell(r,col).border=B; ws4.cell(r,col).alignment=Alignment(horizontal="center")
    r+=1
last_site_row=r-1
for col,w in zip("ABCDEFGHI",[7,9,9,11,11,10,8,9,12]): ws4.column_dimensions[col].width=w
ws4.column_dimensions["K"].width=10; ws4.column_dimensions["L"].width=9
ws4.freeze_panes="A5"

# live wafer-stat block on Site_Detail, computed over INCLUDED sites (F="Y")
sr=last_site_row+2
ws4.cell(sr,1,f"W{REP:02d} statistics (included sites, live)").font=BF
ws4.merge_cells(start_row=sr,start_column=1,end_row=sr,end_column=4)
repx=lot["wafers"][REP-1]
F_rng=f"F5:F{last_site_row}"; I_rng=f"I5:I{last_site_row}"
# MIN/MAX/AVERAGE over the included-only helper column I (blanks ignored).
# These are legacy functions, so no _xlfn prefix issue and they work everywhere.
stat_defs=[
 ("Included site count", f'=COUNTIF({F_rng},"Y")', repx["n"], "int", 0, "0"),
 ("Mean Rs", f'=ROUND(AVERAGE({I_rng}),2)', repx["mean"], "num", 2, "0.00"),
 ("Min Rs", f'=ROUND(MIN({I_rng}),2)', repx["min"], "num", 2, "0.00"),
 ("Max Rs", f'=ROUND(MAX({I_rng}),2)', repx["max"], "num", 2, "0.00"),
]
for i,(lab,formula,cached,kind,dp,fmt) in enumerate(stat_defs):
    rr=sr+1+i
    ws4.cell(rr,1,lab).font=CF
    live.set(ws4, f"C{rr}", formula, cached, kind=kind, dp=dp, number_format=fmt,
             font=MONO, align=Alignment(horizontal="center"))
    for col in (1,2,3): ws4.cell(rr,col).border=B
# WIW NU from the min/max/mean cells above (live, references the stat cells)
mean_c=f"C{sr+2}"; min_c=f"C{sr+3}"; max_c=f"C{sr+4}"
rr=sr+5
ws4.cell(rr,1,"WIW NU %").font=CF
live.set(ws4, f"C{rr}", f"=ROUND(({max_c}-{min_c})/(2*{mean_c})*100,2)", repx["wiw"],
         kind="num", dp=2, number_format="0.00", font=MONO, align=Alignment(horizontal="center"))
for col in (1,2,3): ws4.cell(rr,col).border=B
# yield (live)
rr=sr+6
ws4.cell(rr,1,"Yield % (in spec / included)").font=CF
live.set(ws4, f"C{rr}", f'=ROUND(COUNTIF({"H5:H"+str(last_site_row)},"Y")/COUNTIF({F_rng},"Y")*100,1)',
         repx["yield_pct"], kind="num", dp=1, number_format="0.0", font=MONO,
         align=Alignment(horizontal="center"))
for col in (1,2,3): ws4.cell(rr,col).border=B

wb.properties.creator="Metrology"; wb.properties.title="Lot L7734-02 Wafer Analysis"; wb.properties.lastModifiedBy="Metrology"
out="golden/L7734-02_Wafer_Analysis.xlsx"; wb.save(out)
nc,nf=live.inject(out)
normalize_decimals(out)   # strip binary float tails from all stored values
set_excel_fingerprint(out)
print("saved",out,f"| live cells {nc} across {nf} sheet(s)")
print(f"lot yield {lot['lot_yield']:.2f}%  W2W {lot['w2w_pct']:.2f}%  max WIW {max(x['wiw'] for x in lot['wafers']):.2f}%  disposition {overall}")
print(f"wafers with fails: {[x['wafer'] for x in lot['wafers'] if x['n_fail']>0]}")

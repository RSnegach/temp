# -*- coding: utf-8 -*-
"""Inject cached <v> for every formula cell in the golden after the fix-save,
computing each from the deduped input so cache == Excel recalc. Covers
Lot_Summary, Per_Wafer, Site_Detail, Raw_Data, and the new Wafer_Map formulas.
Then strip float tails and set the Excel fingerprint."""
import sys, re, json, zipfile, shutil, statistics
sys.path.insert(0,"scripts")
import wafer as W
from openpyxl import load_workbook

GOLD="golden/L7734-02_Wafer_Analysis.xlsx"
LSL=80.75; USL=89.25; WARN_LO=82.45; WARN_HI=87.55; EXCL_R=72.0

# ---- deduped truth from INPUT (so golden aligns with input) ----
wbi=load_workbook("inputs/L7734-02_Rs_Measurements.xlsx", data_only=True)
ms=wbi["Measurements"]
raw=[]
for r in range(3,ms.max_row+1):
    w=ms.cell(r,1).value
    if not w: continue
    raw.append((w,ms.cell(r,2).value,ms.cell(r,5).value,ms.cell(r,6).value,ms.cell(r,7).value))
best={}
for w,s,rad,rs,ts in raw:
    k=(w,s)
    if k not in best or ts>best[k][4]: best[k]=(w,s,rad,rs,ts)
ded=sorted(best.values(), key=lambda x:(x[0],x[1]))
wafers=sorted({x[0] for x in ded})

def wstats(w):
    inc=[rs for (ww,s,rad,rs,ts) in ded if ww==w and rad<=EXCL_R]
    n=len(inc); mean=round(statistics.mean(inc),2); mn=round(min(inc),2); mx=round(max(inc),2)
    npass=sum(1 for v in inc if LSL<=v<=USL)
    return dict(n=n,mean=mean,mn=mn,mx=mx,wiw=round((mx-mn)/(2*mean)*100,2),
                yld=round(npass/n*100,2),fails=n-npass)
st={w:wstats(w) for w in wafers}

cache={}
# Per_Wafer
for i,w in enumerate(wafers):
    r=3+i; s=st[w]
    cache[f"Per_Wafer!B{r}"]=s["n"]; cache[f"Per_Wafer!C{r}"]=s["mean"]; cache[f"Per_Wafer!D{r}"]=s["mn"]
    cache[f"Per_Wafer!E{r}"]=s["mx"]; cache[f"Per_Wafer!F{r}"]=s["wiw"]; cache[f"Per_Wafer!G{r}"]=s["yld"]; cache[f"Per_Wafer!H{r}"]=s["fails"]
lot=3+len(wafers)
lot_mean=round(statistics.mean(st[w]["mean"] for w in wafers),2)
means=[st[w]["mean"] for w in wafers]
w2w=round((max(means)-min(means))/(2*lot_mean)*100,2)
maxwiw=round(max(st[w]["wiw"] for w in wafers),2)
totn=sum(st[w]["n"] for w in wafers); totf=sum(st[w]["fails"] for w in wafers)
lotyld=round((totn-totf)/totn*100,2)
cache[f"Per_Wafer!C{lot}"]=lot_mean; cache[f"Per_Wafer!F{lot}"]=maxwiw; cache[f"Per_Wafer!G{lot}"]=lotyld
# Lot_Summary
cache["Lot_Summary!B5"]=lot_mean; cache["Lot_Summary!B6"]=w2w; cache["Lot_Summary!B7"]=maxwiw
cache["Lot_Summary!B8"]=lotyld; cache["Lot_Summary!B9"]=totn; cache["Lot_Summary!B10"]=totf
cache["Lot_Summary!C6"]="PASS" if w2w<=3 else "FAIL"
cache["Lot_Summary!C7"]="PASS" if maxwiw<=5 else "FAIL"
cache["Lot_Summary!C8"]="PASS" if lotyld>=95 else "FAIL"
cache["Lot_Summary!B12"]="CONTINUE" if (w2w<=3 and maxwiw<=5 and lotyld>=95) else "HOLD"

# Site_Detail (rep wafer W01) + Raw_Data recompute from engine matching build
wb=load_workbook(GOLD)
# Site_Detail: rows 3..51 = sites 1..49 of W01
off=W.wafer_offset(1)
for s in W.SITES:
    r=2+s["site"]; rs=W.rs_value(s,1,off); rad=s["r"]; incl=not s["excluded"]
    cache[f"Site_Detail!D{r}"]=round(rad,2)
    cache[f"Site_Detail!F{r}"]="Y" if incl else "N"
    if incl:
        b="FAIL" if (rs<LSL or rs>USL) else ("WARN" if (rs<WARN_LO or rs>WARN_HI) else "PASS")
        cache[f"Site_Detail!G{r}"]=b
        cache[f"Site_Detail!H{r}"]="Y" if LSL<=rs<=USL else "N"
        cache[f"Site_Detail!I{r}"]=round(rs,2)
    else:
        cache[f"Site_Detail!G{r}"]="-"; cache[f"Site_Detail!H{r}"]="-"; cache[f"Site_Detail!I{r}"]=""
# Raw_Data included + Rs_incl
rd=wb["Raw_Data"]
for r in range(3, rd.max_row+1):
    w=rd.cell(r,1).value
    if not w: continue
    rad=rd.cell(r,3).value; rs=rd.cell(r,4).value; inc="Y" if rad<=EXCL_R else "N"
    cache[f"Raw_Data!E{r}"]=inc; cache[f"Raw_Data!F{r}"]=(rs if inc=="Y" else "")
# Wafer_Map: cell value = Site_Detail Rs or "EX"
cx=8;cy=8;sc=6.2/W.WAFER_RADIUS_MM
for s in W.SITES:
    col=1+int(round(cx+s["x"]*sc)); row=3+int(round(cy-s["y"]*sc))
    coord=wb["Wafer_Map"].cell(row,col).coordinate
    if s["excluded"]: cache[f"Wafer_Map!{coord}"]="EX"
    else: cache[f"Wafer_Map!{coord}"]=round(W.rs_value(s,1,off),2)

# ---- inject ----
def sheet_files(z):
    wbx=z.read("xl/workbook.xml").decode(); rels=z.read("xl/_rels/workbook.xml.rels").decode()
    nr={}
    for tag in re.findall(r"<sheet\b[^>]*?/?>",wbx):
        nm=re.search(r'\bname="([^"]+)"',tag); rid=re.search(r'r:id="([^"]+)"',tag)
        if nm and rid: nr[nm.group(1)]=rid.group(1)
    rt={}
    for tag in re.findall(r"<Relationship\b[^>]*?/?>",rels):
        i=re.search(r'\bId="([^"]+)"',tag); t=re.search(r'\bTarget="([^"]+)"',tag)
        if i and t: rt[i.group(1)]=t.group(1)
    o={}
    for n,rid in nr.items():
        tg=rt.get(rid,""); tg=tg[1:] if tg.startswith("/") else ("xl/"+tg.lstrip("./") if not tg.startswith("xl/") else tg)
        o[n]=tg
    return o
def esc(s): return str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;").replace('"',"&quot;")
def emit(v):
    if isinstance(v,int) or (isinstance(v,float) and v==int(v)): return str(int(v)),False
    if isinstance(v,float): return f"{v:.4f}".rstrip("0").rstrip("."),False
    return esc(v),True

zin=zipfile.ZipFile(GOLD); t2f=sheet_files(zin)
perfile={}
for key,val in cache.items():
    sh,co=key.split("!"); perfile.setdefault(t2f[sh],{})[co]=val
def patch(txt,cmap):
    def repl(m):
        cell=m.group(0); cm=re.search(r'\br="([A-Z]+\d+)"',cell)
        if not cm or cm.group(1) not in cmap: return cell
        if "<f>" not in cell and "<f " not in cell: return cell
        val=cmap[cm.group(1)]; cell=re.sub(r"<v>.*?</v>","",cell,flags=re.S)
        s,isstr=emit(val)
        if isstr:
            ot=re.match(r"<c\b[^>]*>",cell).group(0)
            if ' t="' not in ot: cell=cell.replace(ot,ot[:-1]+' t="str">',1)
        return cell.replace("</c>",f"<v>{s}</v></c>",1)
    return re.sub(r"<c\b[^>]*>.*?</c>",repl,txt,flags=re.S)
tmp=GOLD+".t"; out=zipfile.ZipFile(tmp,"w",zipfile.ZIP_DEFLATED); pat=0
for it in zin.infolist():
    d=zin.read(it.filename)
    if it.filename in perfile: d=patch(d.decode(),perfile[it.filename]).encode(); pat+=1
    out.writestr(it,d)
out.close(); zin.close(); shutil.move(tmp,GOLD)
print(f"injected {len(cache)} cached values across {pat} sheets")

# tail strip + fingerprint (both files)
from xlsx_live import normalize_decimals, set_excel_fingerprint
for fp in [GOLD, "inputs/L7734-02_Rs_Measurements.xlsx"]:
    normalize_decimals(fp); set_excel_fingerprint(fp)
print("stripped tails + fingerprint on both files")
